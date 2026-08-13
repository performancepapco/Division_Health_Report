"""Shared helpers used by every section plugin: name normalization, month-key
conversion between the CLI's ISO form (2026-06) and the dashboard's existing
label form (Jun-26, matches BOOKING_BY_MONTH / POSB_BY_MONTH keys already
baked into index.html so Phase 2's loader doesn't have to reshape anything),
and the master office roster loader.
"""
import re
from pathlib import Path
from datetime import date
from collections import defaultdict
import openpyxl

BASE = Path(__file__).parent.parent

_MONTH_ABBR = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
               "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def short_div(s):
    # Handles "... Division" and the "... Divison" typo present in some
    # source rows (e.g. "RMS TP Divison"), so every unit normalizes the same.
    return re.sub(r"\s+Divis(?:i)?on\s*$", "", str(s)).strip() if s else ""


def short_region(s):
    return str(s).replace(" Region", "").strip() if s else ""


def iso_to_label(iso_month: str) -> str:
    """'2026-06' -> 'Jun-26'"""
    y, m = iso_month.split("-")
    return f"{_MONTH_ABBR[int(m) - 1]}-{y[2:]}"


def label_to_iso(label: str) -> str:
    """'Jun-26' -> '2026-06'"""
    mon, yy = label.split("-")
    m = _MONTH_ABBR.index(mon) + 1
    return f"20{yy}-{m:02d}"


def prev_iso_month(iso_month: str) -> str:
    y, m = (int(x) for x in iso_month.split("-"))
    y, m = (y - 1, 12) if m == 1 else (y, m - 1)
    return f"{y:04d}-{m:02d}"


def validate_iso_month(iso_month: str) -> None:
    if not re.fullmatch(r"\d{4}-(0[1-9]|1[0-2])", iso_month):
        raise ValueError(f"month must be YYYY-MM, got {iso_month!r}")


class Roster:
    """Master office roster (Hierarchy_data.xlsx) — not a monthly upload,
    see pipeline_config.yaml circle.master_roster_file. Cached per-process
    since several sections consult it for the same run."""

    _cache = None

    def __init__(self, ids_by_div, ids_by_reg, div_of_office, region_of_office,
                 name_of_office, count_by_div, count_by_reg, offices):
        self.ids_by_div = ids_by_div
        self.ids_by_reg = ids_by_reg
        self.div_of_office = div_of_office
        self.region_of_office = region_of_office
        self.name_of_office = name_of_office
        self.count_by_div = count_by_div
        self.count_by_reg = count_by_reg
        # oid -> {name, type, division, sub_division, region} — full record,
        # for derived views (subdiv/BO-lookup) that need more than counts.
        self.offices = offices

    @classmethod
    def load(cls, cfg: dict) -> "Roster":
        if cls._cache is not None:
            return cls._cache
        circle = cfg["circle"]
        path = BASE / circle["master_roster_file"]
        sheet = circle["master_roster_sheet"]
        real_types = set(circle["real_office_types"])

        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb[sheet]
        rows = ws.iter_rows(values_only=True)
        header = next(rows)
        idx = {h: i for i, h in enumerate(header)}

        ids_by_div, ids_by_reg = defaultdict(set), defaultdict(set)
        div_of_office, region_of_office, name_of_office = {}, {}, {}
        offices = {}
        for r in rows:
            otype = r[idx["office_type_code"]]
            if otype not in real_types:
                continue
            div = short_div(r[idx["division_name"]])
            reg = short_region(r[idx["region_name"]])
            oid = str(r[idx["office_id"]])
            ids_by_div[div].add(oid)
            ids_by_reg[reg].add(oid)
            div_of_office[oid] = div
            region_of_office[oid] = reg
            name = r[idx["office_name"]] if "office_name" in idx else None
            if name is not None:
                name_of_office[oid] = name
            offices[oid] = {
                "name": name,
                "type": otype,
                "division": div,
                "sub_division": (r[idx["sub_division_name"]] if "sub_division_name" in idx else None) or "(Unmapped)",
                "region": reg,
            }

        count_by_div = {d: len(ids) for d, ids in ids_by_div.items()}
        count_by_reg = {r_: len(ids) for r_, ids in ids_by_reg.items()}

        cls._cache = cls(dict(ids_by_div), dict(ids_by_reg), div_of_office,
                          region_of_office, name_of_office, count_by_div, count_by_reg, offices)
        return cls._cache

    @classmethod
    def offices_in_divisions(cls, cfg: dict, real_divisions: set) -> dict:
        """Roster.load() filters by real_office_types only; the legacy
        subdiv/BO-lookup build additionally restricted to divisions that
        actually appeared in that run's POSB data (excludes RMS/admin
        divisions and any stray blank division rows). Returns a fresh
        oid -> record dict filtered that way, without touching the cache."""
        roster = cls.load(cfg)
        return {oid: rec for oid, rec in roster.offices.items() if rec["division"] in real_divisions}


def _oid_key(v):
    """office_id shows up as int in some source files and float (11100001.0)
    in others; normalize both to the same plain-digit string Roster uses."""
    if v is None:
        return None
    return str(int(v)) if isinstance(v, float) else str(v)


class OfficeGeo:
    """Static per-office location reference — pincode, latitude/longitude,
    district, constituency, tribal-area flag. Split across two source files
    (see pipeline_config.yaml circle.office_geo_file / office_district_file),
    both keyed by office_id, merged here since neither file alone has every
    field. Cached per-process like Roster, since this is roster-like
    reference data, not monthly activity."""

    _cache = None

    @classmethod
    def load(cls, cfg: dict) -> dict:
        if cls._cache is not None:
            return cls._cache
        circle = cfg["circle"]
        geo: dict[str, dict] = {}

        def _blank():
            return {"pincode": None, "lat": None, "lon": None,
                     "district": None, "constituency": None, "tribal": False}

        geo_path = BASE / circle["office_geo_file"]
        if geo_path.exists():
            wb = openpyxl.load_workbook(geo_path, data_only=True, read_only=True)
            ws = wb[circle["office_geo_sheet"]]
            rows = ws.iter_rows(values_only=True)
            idx = {h: i for i, h in enumerate(next(rows))}
            for r in rows:
                oid = _oid_key(r[idx["office_id"]])
                if not oid:
                    continue
                rec = geo.setdefault(oid, _blank())
                pincode = r[idx["pincode"]]
                rec["pincode"] = int(pincode) if isinstance(pincode, float) else pincode
                rec["lat"] = r[idx["latitude"]]
                rec["lon"] = r[idx["longitude"]]

        dist_path = BASE / circle["office_district_file"]
        if dist_path.exists():
            wb = openpyxl.load_workbook(dist_path, data_only=True, read_only=True)
            ws = wb[circle["office_district_sheet"]]
            rows = ws.iter_rows(values_only=True)
            idx = {h: i for i, h in enumerate(next(rows))}
            for r in rows:
                oid = _oid_key(r[idx["office_id"]])
                if not oid:
                    continue
                rec = geo.setdefault(oid, _blank())
                rec["district"] = r[idx["District Name"]]
                rec["constituency"] = r[idx["Lok Sabha/Parliamentary/Constituency Name"]]
                tribal = r[idx["Tribal area or not (YES/NO)"]]
                rec["tribal"] = bool(tribal) and str(tribal).strip().upper() == "YES"
                if rec["pincode"] is None:
                    pincode = r[idx["pincode"]]
                    rec["pincode"] = int(pincode) if isinstance(pincode, float) else pincode

        cls._cache = geo
        return cls._cache

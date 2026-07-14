"""ECR — Revenue & Expenditure.

The uploaded file already reports cumulative-to-date figures (the source
system computes the running total), so extract() for this section IS the
cumulative — there is no cross-month summation like POSB/PLI/RPLI/Booking.
See cumulative_mode: source in pipeline_config.yaml.

Column positions are read from pipeline_config.yaml (not hardcoded), and are
only trusted after validate() has confirmed the header row still matches —
so a reordered column fails loudly in CI instead of silently reading the
wrong values, which is what the original per-month script did (fixed
COL_* indices with no header check at all).
"""
import re
import openpyxl

from pipeline.validate import validate_xlsx_sheet, ValidationError
from pipeline.common import iso_to_label

CR = 10_000_000


def _col_map(sheet_cfg: dict) -> dict:
    return {spec["name"]: spec["column"] for spec in sheet_cfg["required_headers"] if "column" in spec}


def _swap_col_map(sheet_cfg: dict, header_row) -> dict:
    header = [str(h).strip() if h is not None else None for h in header_row]
    out = {}
    for spec in sheet_cfg["required_headers"]:
        candidates = {c.strip() for c in [spec["name"]] + spec.get("alternates", [])}
        idx = next((i for i, h in enumerate(header) if h in candidates), None)
        if idx is not None:
            out[spec["name"]] = idx
    return out


def _to_cr(v):
    return round((v or 0) / CR, 7)


def validate(section_cfg: dict, path, month_iso: str) -> list[ValidationError]:
    errors = []
    for sheet_name, sheet_cfg in section_cfg["sheets"].items():
        errors += validate_xlsx_sheet("ecr", "ECR file", path, sheet_name, sheet_cfg)
    if errors:
        return errors  # can't safely read cell positions if headers didn't match

    # Month-in-file check: the "Total rev on sbcc up to <Month>-<Year>" header
    # literally encodes the reporting month — cross-check it against --month.
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["ECR"]
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_map = _col_map(section_cfg["sheets"]["ECR"])
    rev_header = str(header[col_map["Total rev on sbcc up to <current month>"]])
    # Header text uses full month name + 4-digit year, e.g. "June-2026".
    m = re.search(r"up to\s+(\w+)-(\d{4})", rev_header)
    if m:
        full_month, year = m.group(1), m.group(2)
        declared_year = month_iso[:4]
        # Compare using the month abbreviation everyone else uses, tolerant of
        # the header spelling out the full month name.
        declared_abbr = iso_to_label(month_iso).split("-")[0]
        if not full_month.startswith(declared_abbr) or year != declared_year:
            errors.append(ValidationError("ecr",
                f"ECR file: header says 'up to {full_month}-{year}' but you declared "
                f"{iso_to_label(month_iso)} in the Form — please upload the ECR file "
                f"for the month you selected."))
    return errors


def _extract_circle_full(section_cfg: dict, path, all_rows: set) -> dict:
    """CIRCLE_FULL — the circle-wide consolidated position. Sums the SAME
    ECR file over known_divisions + circle_full_extra_rows (33 + 6 = 39
    rows) instead of per-division, matching
    patch_june2026_circle_full.py exactly (including its round(v, 4), not
    round(v, 7) — the two constants were always rounded differently)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ecr_ws = wb["ECR"]
    swap_ws = wb["SWAP"]
    ecr_cfg = section_cfg["sheets"]["ECR"]
    swap_cfg = section_cfg["sheets"]["SWAP"]
    col = _col_map(ecr_cfg)

    actuals = {"Parcel": 0.0, "Mail Operations": 0.0, "IR & GB": 0.0, "CCS": 0.0,
               "POSB": 0.0, "PLI/RPLI": 0.0, "Total": 0.0}
    total_exp = 0.0
    for row in ecr_ws.iter_rows(min_row=ecr_cfg["header_row"] + 1, values_only=True):
        name = row[col["Division"]]
        if not name or name not in all_rows:
            continue
        actuals["Parcel"] += (row[col["Parcel"]] or 0) / CR
        actuals["Mail Operations"] += (row[col["MO"]] or 0) / CR
        actuals["IR & GB"] += (row[col["IRGB"]] or 0) / CR
        actuals["CCS"] += (row[col["CCS"]] or 0) / CR
        actuals["POSB"] += (row[col["Total rev on sbcc up to <current month>"]] or 0) / CR
        actuals["PLI/RPLI"] += (row[col["Total PLI/RPLI Rev."]] or 0) / CR
        actuals["Total"] += (row[col["Total revenue"]] or 0) / CR
        total_exp += (row[col["Exp. For ECR calculation"]] or 0) / CR

    swap_header_row = next(swap_ws.iter_rows(min_row=1, max_row=1, values_only=True))
    swap_col = _swap_col_map(swap_cfg, swap_header_row)
    other_key = "Others" if "Others" in swap_col else None
    swap = {"Salaries": 0.0, "Wages": 0.0, "Pension": 0.0, "Allowances": 0.0, "Plan": 0.0, "Other": 0.0}
    for row in swap_ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        if not name or name not in all_rows:
            continue
        swap["Salaries"] += (row[swap_col["Salaries"]] or 0) / CR
        swap["Wages"] += (row[swap_col["Wages"]] or 0) / CR
        swap["Pension"] += (row[swap_col["Pension"]] or 0) / CR
        swap["Allowances"] += (row[swap_col["Allowances"]] or 0) / CR
        swap["Plan"] += (row[swap_col["Plan Expnditure "]] or 0) / CR
        if other_key:
            swap["Other"] += (row[swap_col[other_key]] or 0) / CR
    swap["Total"] = sum(swap.values())

    ecr_with_pension = (actuals["Total"] / total_exp * 100) if total_exp > 0 else 0
    denom = total_exp - swap["Pension"]
    ecr_without_pension = (actuals["Total"] / denom * 100) if denom > 0 else 0

    return {
        "actuals": {k: round(v, 4) for k, v in actuals.items()},
        "total_exp": round(total_exp, 4),
        "swap": {k: round(v, 4) for k, v in swap.items()},
        "ecr_with_pension": round(ecr_with_pension, 4),
        "ecr_without_pension": round(ecr_without_pension, 4),
    }


def extract(section_cfg: dict, path, month_iso: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    ecr_ws = wb["ECR"]
    swap_ws = wb["SWAP"]
    ecr_cfg = section_cfg["sheets"]["ECR"]
    swap_cfg = section_cfg["sheets"]["SWAP"]
    col = _col_map(ecr_cfg)
    known_divisions = set(section_cfg["known_divisions"])
    # Circle/region/PSD residual rows (e.g. "AP GM (Finance)") are extracted
    # alongside the 33 known divisions, keyed by the same row names the
    # front end's OFFICE_ENTRIES/applyEcrMonth patcher expects — so those
    # residuals get refreshed every month same as any division, instead of
    # staying frozen at their hand-authored April baseline forever. See
    # circle_full_extra_rows in pipeline_config.yaml.
    extra_rows = set(section_cfg["circle_full_extra_rows"])
    all_rows = known_divisions | extra_rows

    out = {}
    for row in ecr_ws.iter_rows(min_row=ecr_cfg["header_row"] + 1, values_only=True):
        name = row[col["Division"]]
        if not name or name not in all_rows:
            continue
        out[name] = {
            "actuals": {
                "Parcel":          _to_cr(row[col["Parcel"]]),
                "Mail Operations": _to_cr(row[col["MO"]]),
                "IR & GB":         _to_cr(row[col["IRGB"]]),
                "CCS":             _to_cr(row[col["CCS"]]),
                "POSB":            _to_cr(row[col["Total rev on sbcc up to <current month>"]]),
                "PLI/RPLI":        _to_cr(row[col["Total PLI/RPLI Rev."]]),
                "Total":           _to_cr(row[col["Total revenue"]]),
            },
            "total_exp": _to_cr(row[col["Exp. For ECR calculation"]]),
        }

    swap_header_row = next(swap_ws.iter_rows(min_row=1, max_row=1, values_only=True))
    swap_col = _swap_col_map(swap_cfg, swap_header_row)
    other_key = "Others" if "Others" in swap_col else None
    for row in swap_ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        if not name or name not in all_rows or name not in out:
            continue
        out[name]["swap"] = {
            "Salaries":   _to_cr(row[swap_col["Salaries"]]),
            "Wages":      _to_cr(row[swap_col["Wages"]]),
            "Pension":    _to_cr(row[swap_col["Pension"]]),
            "Allowances": _to_cr(row[swap_col["Allowances"]]),
            "Plan":       _to_cr(row[swap_col["Plan Expnditure "]]),
            "Other":      _to_cr(row[swap_col[other_key]]) if other_key else 0,
            "Total":      _to_cr(row[swap_col["Total expenditure"]]),
        }

    missing = known_divisions - set(out.keys())
    if missing:
        print(f"  WARNING: missing in Excel: {sorted(missing)}")

    circle_full = _extract_circle_full(section_cfg, path, all_rows)

    return {"divisions": out, "circle_full": circle_full}


def merge_cumulative(new_slice: dict, previous_cumulative: dict | None) -> dict:
    # cumulative_mode: source — the file already IS the running total.
    return new_slice


def row_count(data: dict) -> int:
    return len(data.get("divisions", {})) if data else 0

"""POSB — Accounts Opened/Closed.

Standalone-month activity sheet: offices with zero activity are simply
absent from it, which understates "offices" per division. The true office
universe comes from the master roster (see pipeline.common.Roster), unioned
with this month's activity IDs (an office with real activity but missing
from the roster — "hierarchy drift" — is still a real office; folded in
rather than dropped).

Note: unlike the legacy build_posb_june2026.py (which pooled activity IDs
across all 3 hardcoded months before computing office universes), this
computes the union per-month since each run only sees one month's file.
A division's "offices" figure may shift very slightly run-to-run if a
drifted office's first appearance moves between months — it converges once
the roster is updated, and does not affect opened/closed/net (a NIL office
contributes 0 to those regardless).

No cross-month merge happens here (cumulative_mode: none) — the dashboard
sums opened/closed/net across the selected month range client-side
(getPOSBData in index.html); see Phase 2 notes.
"""
import re
from collections import defaultdict
import openpyxl

from pipeline.validate import validate_xlsx_sheet, ValidationError
from pipeline.common import short_div, short_region, Roster, iso_to_label
from pipeline.config import load_config

SHEET = "2. BO Totals"
SCHEME_SHEET = "1. Scheme-wise Detail"
POSB_SCHEMES = ["MIS", "PPFGP", "SSA", "RD", "SBBAS", "SBSGP", "SCSS", "TD",
                "PRFTS", "KVN", "NSC8", "MSSC"]


def _header_map(ws) -> dict:
    """name -> column index, read from row 1. Column *positions* aren't
    trustworthy across sources — e.g. the uploader's own POSB export has an
    extra "Office Type" column the Directorate's template doesn't — so every
    reader in this module looks columns up by header name instead of a
    hardcoded index."""
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return {str(h).strip(): i for i, h in enumerate(header) if h is not None}


def validate(section_cfg: dict, path, month_iso: str) -> list[ValidationError]:
    sheet_cfg = section_cfg["sheets"][SHEET]
    errors = validate_xlsx_sheet("posb", "POSB file", path, SHEET, sheet_cfg)
    if errors:
        return errors

    # Month-in-file check: the "Month" column holds a date range like
    # "01-06-2026 - 30-06-2026" (DD-MM-YYYY). Compare its month/year against
    # the declared month for the first data row (all rows share one range).
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[SHEET]
    month_idx = _header_map(ws)["Month"]
    rows = ws.iter_rows(min_row=2, values_only=True)
    first = next(rows, None)
    if first is None:
        return [ValidationError("posb", "POSB file: no data rows found below the header.")]
    month_col = first[month_idx]
    m = re.search(r"-(\d{2})-(\d{4})", str(month_col) or "")
    if m:
        file_month, file_year = m.group(1), m.group(2)
        declared_month = month_iso[5:7]
        declared_year = month_iso[:4]
        if file_month != declared_month or file_year != declared_year:
            return [ValidationError("posb",
                f"POSB file: 'Month' column says {month_col!r} but you declared "
                f"{iso_to_label(month_iso)} in the Form — please upload the POSB "
                f"file for the month you selected.")]
    return []


def _read_schemes(wb) -> dict:
    """'1. Scheme-wise Detail' — per-office scheme breakdown, consumed only
    by the derived BO_LOOKUP view (pipeline/derive/subdiv_bolookup.py), not
    by anything in this section's own by_div/by_reg/circle rollups. Kept
    positional (base = 7 + i*3 per scheme), matching the source report's
    lack of stable per-scheme headers — same approach the legacy
    build_subdiv_bolookup_june2026.py used."""
    if SCHEME_SHEET not in wb.sheetnames:
        return {}
    ws = wb[SCHEME_SHEET]
    hmap = _header_map(ws)
    oid_idx, code_idx = hmap["Office ID"], hmap["BO Code"]
    scheme_cols = [(hmap[f"{name} Opened"], hmap[f"{name} Closed"], hmap[f"Net {name}"])
                   for name in POSB_SCHEMES]
    schemes = defaultdict(dict)
    for row in ws.iter_rows(min_row=2, values_only=True):
        oid, code = row[oid_idx], row[code_idx]
        if not oid and not code:
            continue
        oid = str(oid) if oid else f"BOCODE_{code}"
        sch = {}
        for name, (oi, ci, ni) in zip(POSB_SCHEMES, scheme_cols):
            o, c, n = int(row[oi] or 0), int(row[ci] or 0), int(row[ni] or 0)
            if o == 0 and c == 0 and n == 0:
                continue
            sch[name] = {"o": o, "c": c, "n": n}
        if sch:
            schemes[oid] = sch
    return dict(schemes)


def extract(section_cfg: dict, path, month_iso: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[SHEET]
    hmap = _header_map(ws)
    oid_idx = hmap["Office ID"]
    code_idx = hmap["BO Code"]
    name_idx = hmap["Name of the BO"]
    sub_div_idx = hmap["Name of the Sub Division"]
    div_idx = hmap["Name of the Division"]
    opened_idx = hmap["Accounts Opened"]
    closed_idx = hmap["Accounts Closed"]
    net_idx = hmap["Net Addition of the Accounts"]
    region_idx = hmap["Region"]

    offices = {}
    by_div = defaultdict(lambda: {"region": "", "offices": 0, "opened": 0, "closed": 0, "net": 0})
    by_reg = defaultdict(lambda: {"offices": 0, "opened": 0, "closed": 0, "net": 0})
    circle = {"offices": 0, "opened": 0, "closed": 0, "net": 0}

    for row in ws.iter_rows(min_row=2, values_only=True):
        oid = row[oid_idx]
        code_fallback = row[code_idx]
        if not oid and not code_fallback:
            continue
        oid = str(oid) if oid else f"BOCODE_{code_fallback}"
        code, name, sub_div, div, opened, closed, net, region = (
            row[code_idx], row[name_idx], row[sub_div_idx], row[div_idx],
            row[opened_idx] or 0, row[closed_idx] or 0, row[net_idx] or 0, row[region_idx] or ""
        )
        div_s = short_div(div)
        reg_s = short_region(region)
        opened, closed, net = int(opened), int(closed), int(net)

        offices[oid] = {
            "code": code, "name": name, "type": "",
            "sub_division": sub_div or "(Unmapped)", "division": div_s, "region": reg_s,
            "opened": opened, "closed": closed, "net": net,
        }

        d = by_div[div_s]
        d["region"] = reg_s
        d["offices"] += 1
        d["opened"] += opened
        d["closed"] += closed
        d["net"] += net

        r = by_reg[reg_s]
        r["offices"] += 1
        r["opened"] += opened
        r["closed"] += closed
        r["net"] += net

        circle["offices"] += 1
        circle["opened"] += opened
        circle["closed"] += closed
        circle["net"] += net

    roster = Roster.load(load_config())
    activity_ids_by_div = defaultdict(set)
    activity_ids_by_reg = defaultdict(set)
    for oid, rec in offices.items():
        activity_ids_by_div[rec["division"]].add(oid)
        activity_ids_by_reg[rec["region"]].add(oid)

    real_divisions = set(by_div.keys())
    final_ids_by_div = {
        div: roster.ids_by_div.get(div, set()) | activity_ids_by_div.get(div, set())
        for div in real_divisions
    }
    final_ids_by_reg = {
        reg: roster.ids_by_reg.get(reg, set()) | activity_ids_by_reg.get(reg, set())
        for reg in set(roster.ids_by_reg) | set(activity_ids_by_reg)
    }
    master_div_counts = {div: len(ids) for div, ids in final_ids_by_div.items()}
    master_reg_counts = {reg: len(ids) for reg, ids in final_ids_by_reg.items()}
    master_circle_count = len(set().union(*final_ids_by_div.values())) if final_ids_by_div else 0

    month_data = {}
    for div, d in by_div.items():
        month_data[div] = {
            "offices": master_div_counts.get(div, d["offices"]), "opened": d["opened"],
            "closed": d["closed"], "net": d["net"], "region": d["region"],
        }
    for reg, r in by_reg.items():
        month_data[f"__REG_{reg}__"] = {
            "offices": master_reg_counts.get(reg, r["offices"]), "opened": r["opened"],
            "closed": r["closed"], "net": r["net"],
        }
    month_data["__CIRCLE__"] = dict(circle, offices=master_circle_count)

    schemes = _read_schemes(wb)

    # "Month" column holds a date range like "01-07-2026 - 27-07-2026" (see
    # validate() above) — its end date is the real as-of date for this
    # file, which for a daily_reupload month-to-date upload is today's
    # pull, not the calendar month-end. Surfaced via subdiv_bolookup.py so
    # BO Lookup can show a real date instead of a stale quarter label.
    as_of = None
    first_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), None)
    if first_row is not None:
        m = re.search(r"(\d{2})-(\d{2})-(\d{4})\s*$", str(first_row[hmap["Month"]] or ""))
        if m:
            dd, mm, yyyy = m.groups()
            as_of = f"{yyyy}-{mm}-{dd}"

    return {"divisions": month_data, "offices": offices, "schemes": schemes, "as_of": as_of}


def merge_cumulative(new_slice: dict, previous_cumulative: dict | None) -> dict:
    # cumulative_mode: none — the dashboard sums opened/closed/net across the
    # selected month range client-side; nothing to merge server-side.
    return new_slice


def row_count(data: dict) -> int:
    return len(data.get("offices", {})) if data else 0

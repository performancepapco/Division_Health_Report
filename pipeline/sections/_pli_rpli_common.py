"""Shared extraction logic for PLI and RPLI — identical file shape, two
separate uploads/sections (pli.py, rpli.py are thin wrappers around this).

No cross-month merge happens here either (cumulative_mode: none) for
policies/premium — the dashboard sums those client-side across the selected
month range (getInsData in index.html). The one field that can't be summed
that way is offices_procured (a distinct-office count, not additive across
months) — the legacy pipeline got that from a separately-uploaded
"cumulative distribution" file. Since the Form intentionally has no such
upload (see Phase 0 section-list confirmation), Phase 2's assembly step will
derive it instead from office-level Detail records unioned across all known
months for the range — deferred there, not needed for per-month extraction.
"""
import re
import openpyxl

from pipeline.validate import validate_xlsx_sheet, ValidationError
from pipeline.common import short_div, iso_to_label

HEADER_RE = re.compile(r"^(.+?)\s+Division\s")


def validate(section_name: str, section_cfg: dict, path, month_iso: str,
             division_sheet: str, detail_sheet: str) -> list[ValidationError]:
    errors = []
    errors += validate_xlsx_sheet(section_name, f"{section_name.upper()} file", path,
                                   division_sheet, section_cfg["sheets"][division_sheet])
    errors += validate_xlsx_sheet(section_name, f"{section_name.upper()} file", path,
                                   detail_sheet, section_cfg["sheets"][detail_sheet])
    if errors:
        return errors

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[division_sheet]
    rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
    title_row = next((r[0] for r in rows if isinstance(r[0], str) and re.match(r"^\w+\s+\d{4}", r[0])), None)
    if title_row:
        m = re.match(r"^(\w+)\s+(\d{4})", title_row)
        full_month, year = m.group(1), m.group(2)
        declared_abbr = iso_to_label(month_iso).split("-")[0]
        declared_year = month_iso[:4]
        if not full_month.startswith(declared_abbr) or year != declared_year:
            errors.append(ValidationError(section_name,
                f"{section_name.upper()} file: title row says '{full_month} {year}' but you "
                f"declared {iso_to_label(month_iso)} in the Form — please upload the "
                f"{section_name.upper()} file for the month you selected."))
    return errors


def _read_division_level(path, sheet):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    hdr_idx = next(i for i, r in enumerate(rows) if r[0] == "Sl.")
    out = {}
    circle = None
    for r in rows[hdr_idx + 1:]:
        if r[1] is None:
            continue
        div_raw = r[1]
        total_off, off_proc = r[3] or 0, r[4] or 0
        policies, premium = r[12] or 0, r[13] or 0
        region = r[2] or ""
        if str(div_raw).strip() == "CIRCLE TOTAL":
            circle = {
                "total_offices": int(total_off), "offices_procured": int(off_proc),
                "policies": int(policies), "premium": float(premium),
            }
            continue
        div = short_div(div_raw)
        procurement_rate = round(off_proc / total_off * 100, 1) if total_off else 0
        avg_premium = round(premium / policies) if policies else 0
        out[div] = {
            "total_offices": int(total_off), "offices_procured": int(off_proc),
            "policies": int(policies), "premium": float(premium),
            "procurement_rate": procurement_rate, "avg_premium": avg_premium,
            "region": str(region).replace(" Region", "").strip(),
        }
    return out, circle


def _read_office_level(path, sheet):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    cur_div = None
    records = []
    unmatched_div_rows = 0
    for row in ws.iter_rows(values_only=True):
        a = row[0]
        if a is None:
            continue
        if isinstance(a, str):
            m = HEADER_RE.match(a)
            if m and "band" in a.lower():
                cur_div = short_div(m.group(1).strip())
                continue
            if a == "Sl.":
                continue
        if not isinstance(row[0], int):
            continue
        if cur_div is None:
            unmatched_div_rows += 1
            continue
        code, name, otype, pols, prem = row[1], row[2], row[3], row[4], row[5]
        if not name:
            continue
        records.append({
            "division": cur_div, "code": code, "name": name, "type": otype or "",
            "policies": int(pols or 0), "premium": float(prem or 0),
        })
    return records, unmatched_div_rows


def extract(path, division_sheet: str, detail_sheet: str) -> dict:
    div_data, circle = _read_division_level(path, division_sheet)
    div_data["__CIRCLE__"] = circle
    offices, unmatched = _read_office_level(path, detail_sheet)
    if unmatched:
        print(f"  NOTE: {unmatched} office row(s) skipped before first division header")
    return {"divisions": div_data, "offices": offices}


def merge_cumulative(new_slice: dict, previous_cumulative: dict | None) -> dict:
    return new_slice


def row_count(data: dict) -> int:
    return len(data.get("offices", [])) if data else 0

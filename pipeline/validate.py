"""Generic, config-driven validation primitives shared by every section.

Header/sheet checks here are driven entirely by pipeline_config.yaml, so a
template and its validation rule can never drift apart (the CI gate and
generate_templates.py both read the same config). Section-specific checks
that can't be expressed generically (e.g. "does the file's own reporting
month match what was declared") live in each pipeline/sections/<name>.py
module instead, and get appended to the same ValidationError list.

Every failure here is meant to be shown, verbatim, to a non-technical
uploader — so messages name the section, the sheet/column, and the exact
problem in plain language. No stack traces, no internal jargon.
"""
import csv
from dataclasses import dataclass
import openpyxl


@dataclass
class ValidationError:
    section: str
    message: str  # plain-language, uploader-facing


def _header_matches(cell_value, spec: dict) -> bool:
    match = spec.get("match", "exact")
    name = spec["name"]
    if match == "exact":
        candidates = {c.strip() for c in [name] + spec.get("alternates", [])}
        return str(cell_value).strip() in candidates if cell_value is not None else False
    if match == "prefix":
        return isinstance(cell_value, str) and cell_value.strip().startswith(spec["prefix"])
    if match == "numeric":
        try:
            return int(cell_value) == int(name)
        except (TypeError, ValueError):
            return False
    raise ValueError(f"unknown header match mode: {match}")


def _find_dynamic_header_row(rows, marker_column=0, marker_value="Sl."):
    for i, row in enumerate(rows):
        if row and row[marker_column] is not None and str(row[marker_column]).strip() == marker_value:
            return i, row
    return None, None


def validate_xlsx_sheet(section_name: str, label: str, path, sheet_name: str,
                         sheet_cfg: dict) -> list[ValidationError]:
    errors = []
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        return [ValidationError(section_name,
            f"{label}: could not open the file — is it a valid Excel (.xlsx) file? ({e})")]

    if sheet_name not in wb.sheetnames:
        return [ValidationError(section_name,
            f"{label}: expected a sheet named '{sheet_name}' — found {wb.sheetnames}. "
            f"Please use the template from /templates and don't rename sheets.")]

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    if sheet_cfg["header_row"] == "dynamic":
        _, header = _find_dynamic_header_row(rows)
        if header is None:
            return [ValidationError(section_name,
                f"{label}, sheet '{sheet_name}': could not find the header row "
                f"(expected a row starting with 'Sl.'). The sheet layout looks different "
                f"from the template — please re-download it from /templates.")]
    else:
        row_num = sheet_cfg["header_row"]
        if len(rows) < row_num:
            return [ValidationError(section_name,
                f"{label}, sheet '{sheet_name}': sheet has fewer than {row_num} rows, "
                f"can't find the header row.")]
        header = rows[row_num - 1]

    for spec in sheet_cfg["required_headers"]:
        col = spec.get("column")
        if col is not None:
            cell = header[col] if col < len(header) else None
            if not _header_matches(cell, spec):
                errors.append(ValidationError(section_name,
                    f"{label}, sheet '{sheet_name}': column {col + 1} should be "
                    f"'{spec['name']}' but found {cell!r}. Please use the template "
                    f"from /templates and don't reorder or rename columns."))
        else:
            candidates = {c.strip() for c in [spec["name"]] + spec.get("alternates", [])}
            if not any(str(h).strip() in candidates for h in header if h is not None):
                errors.append(ValidationError(section_name,
                    f"{label}, sheet '{sheet_name}': missing expected column "
                    f"'{spec['name']}'. Please use the template from /templates."))
    return errors


def validate_csv(section_name: str, label: str, path, csv_cfg: dict) -> list[ValidationError]:
    try:
        with open(path, newline="", encoding="utf-8-sig") as f:
            header = next(csv.reader(f), None)
    except Exception as e:
        return [ValidationError(section_name,
            f"{label}: could not open the file as CSV ({e})")]

    if header is None:
        return [ValidationError(section_name, f"{label}: file is empty.")]

    header_set = {h.strip() for h in header if h is not None}
    missing = [h for h in csv_cfg["required"] if h not in header_set]
    if missing:
        errors = [ValidationError(section_name,
            f"{label}: missing column '{h}' — found columns: {sorted(header_set)}. "
            f"Please use the template from /templates and don't rename columns.")
            for h in missing]
        return errors
    return []


def validate_row_count(section_name: str, label: str, new_count: int,
                        prev_count: int | None, band: list,
                        baseline: str = "last month's") -> list[ValidationError]:
    """band's upper bound (band[1]) may be None to skip the upper check
    entirely — used for same-month re-uploads, where growth through the
    month is expected and not suspicious, only a big drop is."""
    if prev_count is None or prev_count == 0:
        return []
    lo, hi = band
    ratio = new_count / prev_count
    if ratio < lo or (hi is not None and ratio > hi):
        hi_txt = f"{hi:.0%}" if hi is not None else "no upper limit"
        return [ValidationError(section_name,
            f"{label}: row count ({new_count:,}) is {ratio:.0%} of {baseline} "
            f"({prev_count:,}) — outside the expected {lo:.0%}-{hi_txt} range. "
            f"This usually means a partial upload or a wrong file. If this drop/spike "
            f"is genuinely correct, an admin can approve it manually from staging.")]
    return []

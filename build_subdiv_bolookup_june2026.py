#!/usr/bin/env python3
"""
build_subdiv_bolookup_june2026.py
────────────────────────────────────
Rebuilds SUBDIV_DATA and BO_LOOKUP as a true Apr+May+Jun 2026 cumulative
snapshot (matching the established "always latest cumulative" design of
these two tabs — they don't respond to the global period picker).

Sources:
  - POSB: uploads/june_2026/POSB/*.xlsx (Apr, May) + uploads/june_2026/Merged_Product_Wise_AC_Report.xlsx (Jun)
          "1. Scheme-wise Detail" (scheme breakdown) + "2. BO Totals" (opened/closed/net)
  - PLI/RPLI: uploads/june_2026/Insurance/*.xlsx (Apr, May) + uploads/june_2026 root (Jun)
              "Detail" sheets, office-level policies/premium
  - Booking: uploads/{april_2026,may_2026,june_2026}/Booking_Productwise_Report*.csv

POSB and PLI/RPLI use different office-ID schemes with no shared key, so —
matching the existing codebase's approach — offices are joined by
(division, normalized office name).

Writes:
  _subdiv_data_june.json   (division -> subdivision -> {region, offices, posb, pli, rpli, bos[]})
  _bo_lookup_june.json     (office list + search index)

Run:  python build_subdiv_bolookup_june2026.py
"""
import json, re, csv
from pathlib import Path
from collections import defaultdict
import openpyxl

BASE = Path(__file__).parent
UP = BASE / "uploads"

POSB_FILES = {
    "Apr-26": UP / "june_2026" / "POSB" / "Merged_Product_Wise_AC_Report_April2026.xlsx",
    "May-26": UP / "june_2026" / "POSB" / "Merged_Product_Wise_AC_Report_may2026.xlsx",
    "Jun-26": UP / "june_2026" / "Merged_Product_Wise_AC_Report.xlsx",
}
INS = UP / "june_2026" / "Insurance"
JUN = UP / "june_2026"
PLI_FILES = {
    "Apr-26": INS / "PLI_April2026_Distribution.xlsx",
    "May-26": INS / "PLI_May2026_Distribution.xlsx",
    "Jun-26": JUN / "PLI_June2026_Distribution.xlsx",
}
RPLI_FILES = {
    "Apr-26": INS / "RPLI_April2026_Distribution.xlsx",
    "May-26": INS / "RPLI_May2026_Distribution.xlsx",
    "Jun-26": JUN / "RPLI_June2026_Distribution.xlsx",
}
BOOKING_PROD = {
    "Apr-26": UP / "april_2026" / "Booking_Productwise_Report apr 2026.csv",
    "May-26": UP / "may_2026" / "Booking_Productwise_Report (2).csv",
    "Jun-26": UP / "june_2026" / "Booking_Productwise_Report.csv",
}

POSB_SCHEMES = ["MIS", "PPFGP", "SSA", "RD", "SBBAS", "SBSGP", "SCSS", "TD", "PRFTS", "KVN", "NSC8", "MSSC"]


def norm_name(s):
    if not s:
        return ""
    s = str(s).strip().lower()
    s = re.sub(r"\s+", " ", s)
    s = s.replace(".", "").replace(",", "")
    s = re.sub(r"\bb\.?o\b", "bo", s)
    s = re.sub(r"\bs\.?o\b", "so", s)
    s = re.sub(r"\bh\.?o\b", "ho", s)
    return s.strip()


def short_div(s):
    return re.sub(r"\s+Division\s*$", "", str(s)).strip() if s else ""


def short_region(s):
    return str(s).replace(" Region", "").strip() if s else ""


def infer_type(name):
    """New source files dropped the 'Office Type' column; approximate from name suffix."""
    n = str(name or "").strip()
    if re.search(r"\bH\.?O\b\.?$", n, re.IGNORECASE):
        return "HPO"
    if re.search(r"\bS\.?O\b\.?$", n, re.IGNORECASE):
        return "SPO"
    return "BPO"


# ─── 1. POSB office master (BO Totals, summed Apr+May+Jun) + scheme-wise ─────
def read_posb_month(path):
    """Returns office dict keyed by Office ID -> record; plus scheme totals keyed by Office ID."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["2. BO Totals"]
    offices = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        oid, code = row[1], row[2]
        if not oid and not code:
            continue
        oid = str(oid) if oid else f"BOCODE_{code}"
        offices[oid] = {
            "code": code or "", "name": row[4], "type": infer_type(row[4]),
            "sub_division": row[5] or "(Unmapped)", "division": short_div(row[6]),
            "region": short_region(row[10]),
            "opened": int(row[7] or 0), "closed": int(row[8] or 0), "net": int(row[9] or 0),
        }

    schemes = defaultdict(dict)
    ws1 = wb["1. Scheme-wise Detail"]
    for row in ws1.iter_rows(min_row=2, values_only=True):
        oid, code = row[1], row[2]
        if not oid and not code:
            continue
        oid = str(oid) if oid else f"BOCODE_{code}"
        sch = {}
        for i, name in enumerate(POSB_SCHEMES):
            base = 7 + i * 3
            o, c, n = int(row[base] or 0), int(row[base + 1] or 0), int(row[base + 2] or 0)
            if o == 0 and c == 0 and n == 0:
                continue
            sch[name] = {"o": o, "c": c, "n": n}
        if sch:
            schemes[oid] = sch
    return offices, schemes


def build_posb_cumulative():
    cum_offices = {}
    cum_schemes = defaultdict(lambda: defaultdict(lambda: {"o": 0, "c": 0, "n": 0}))
    for month, path in POSB_FILES.items():
        print(f"  POSB {month}: {path.name}")
        offices, schemes = read_posb_month(path)
        for oid, rec in offices.items():
            if oid not in cum_offices:
                cum_offices[oid] = dict(rec, opened=0, closed=0, net=0)
            cum_offices[oid]["opened"] += rec["opened"]
            cum_offices[oid]["closed"] += rec["closed"]
            cum_offices[oid]["net"] += rec["net"]
            # keep latest name/division/subdivision (hierarchy may be refreshed)
            for k in ("name", "code", "division", "sub_division", "region"):
                if rec.get(k):
                    cum_offices[oid][k] = rec[k]
        for oid, sch in schemes.items():
            for scheme_name, v in sch.items():
                acc = cum_schemes[oid][scheme_name]
                acc["o"] += v["o"]; acc["c"] += v["c"]; acc["n"] += v["n"]
    return cum_offices, cum_schemes


# ─── 2. PLI / RPLI office detail (Detail sheet, summed Apr+May+Jun) ──────────
HEADER_RE = re.compile(r"^(.+?)\s+Division\s")


def read_ins_detail_month(path, sheet):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    cur_div = None
    recs = []
    for row in ws.iter_rows(values_only=True):
        a = row[0]
        if a is None:
            continue
        if isinstance(a, str):
            m = HEADER_RE.match(a)
            if m and "band" in a.lower():
                cur_div = short_div(m.group(1).strip())
                continue
            if a == "Sl.":
                continue
        if not isinstance(row[0], int) or cur_div is None:
            continue
        name = row[2]
        if not name:
            continue
        recs.append({"division": cur_div, "name": name, "policies": int(row[4] or 0), "premium": float(row[5] or 0)})
    return recs


def build_ins_cumulative(files, sheet_name):
    cum = defaultdict(lambda: {"policies": 0, "premium": 0.0})
    for month, path in files.items():
        print(f"  {sheet_name} {month}: {path.name}")
        recs = read_ins_detail_month(path, sheet_name)
        for r in recs:
            key = (r["division"], norm_name(r["name"]))
            cum[key]["policies"] += r["policies"]
            cum[key]["premium"] += r["premium"]
    return cum


# ─── 3. Booking per-office per-product, summed Apr+May+Jun ──────────────────
def read_booking_cumulative():
    cum = defaultdict(lambda: defaultdict(lambda: {"articles": 0, "amount": 0.0}))
    for month, path in BOOKING_PROD.items():
        if not path.exists():
            print(f"  WARN: {path} missing, skipping {month}")
            continue
        n = 0
        with open(path, newline="", encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                oid = (row.get("office-id") or "").strip()
                if not oid or oid == "-":
                    continue
                prod = (row.get("product-name") or "").strip()
                if not prod:
                    continue
                arts = int(row.get("article-count") or 0)
                amt = float(row.get("total_amount") or 0)
                cum[oid][prod]["articles"] += arts
                cum[oid][prod]["amount"] += amt
                n += 1
        print(f"  Booking {month}: {n:,} rows")
    return cum


def main():
    print("Building POSB cumulative (office + scheme-wise)...")
    posb_offices, posb_schemes = build_posb_cumulative()
    print(f"  {len(posb_offices):,} offices")

    print("Building PLI cumulative...")
    pli_cum = build_ins_cumulative(PLI_FILES, "PLI Detail")
    print("Building RPLI cumulative...")
    rpli_cum = build_ins_cumulative(RPLI_FILES, "RPLI Detail")

    print("Building booking cumulative...")
    booking_cum = read_booking_cumulative()

    # ── Build SUBDIV_DATA: division -> subdivision -> {region, offices, posb, pli, rpli, bos[]} ──
    print("Assembling SUBDIV_DATA...")
    by_div = defaultdict(lambda: defaultdict(lambda: {
        "region": "", "offices": 0,
        "posb": {"opened": 0, "closed": 0, "net": 0},
        "pli": {"offices_procured": 0, "policies": 0, "premium": 0, "non_procurers": 0},
        "rpli": {"offices_procured": 0, "policies": 0, "premium": 0, "non_procurers": 0},
        "bos": [],
    }))
    bo_lookup = {}
    name_to_oid = {}

    for oid, rec in posb_offices.items():
        div_s, sd = rec["division"], rec["sub_division"]
        pli_v = pli_cum.get((div_s, norm_name(rec["name"])), {"policies": 0, "premium": 0.0})
        rpli_v = rpli_cum.get((div_s, norm_name(rec["name"])), {"policies": 0, "premium": 0.0})

        bo_rec = {
            "id": oid, "code": rec["code"], "name": rec["name"], "type": rec["type"],
            "posb_opened": rec["opened"], "posb_closed": rec["closed"], "posb_net": rec["net"],
            "pli_policies": pli_v["policies"], "pli_premium": pli_v["premium"],
            "rpli_policies": rpli_v["policies"], "rpli_premium": rpli_v["premium"],
        }
        s = by_div[div_s][sd]
        s["region"] = rec["region"]
        s["offices"] += 1
        s["posb"]["opened"] += rec["opened"]
        s["posb"]["closed"] += rec["closed"]
        s["posb"]["net"] += rec["net"]
        s["bos"].append(bo_rec)
        name_to_oid[(div_s, norm_name(rec["name"]))] = oid

        # BO_LOOKUP record
        bk = booking_cum.get(oid, {})
        prods = []
        tot_a, tot_amt = 0, 0.0
        for pname, v in sorted(bk.items(), key=lambda kv: -kv[1]["amount"]):
            if v["articles"] == 0 and v["amount"] == 0:
                continue
            prods.append({"n": pname, "a": v["articles"], "v": round(v["amount"], 2)})
            tot_a += v["articles"]; tot_amt += v["amount"]

        bo_lookup[oid] = {
            "id": oid, "code": rec["code"], "name": rec["name"], "type": rec["type"],
            "sub_division": sd, "division": div_s, "region": rec["region"], "pincode": None,
            "posb": {"opened": rec["opened"], "closed": rec["closed"], "net": rec["net"]},
            "posb_schemes": posb_schemes.get(oid, {}),
            "pli": {"policies": pli_v["policies"], "premium": pli_v["premium"]},
            "rpli": {"policies": rpli_v["policies"], "premium": rpli_v["premium"]},
            "booking": {"total_articles": tot_a, "total_amount": round(tot_amt, 2), "products": prods},
        }

    for div, subs in by_div.items():
        for sd, s in subs.items():
            pli_off = sum(1 for b in s["bos"] if b["pli_policies"] > 0)
            rpli_off = sum(1 for b in s["bos"] if b["rpli_policies"] > 0)
            s["pli"]["offices_procured"] = pli_off
            s["pli"]["policies"] = sum(b["pli_policies"] for b in s["bos"])
            s["pli"]["premium"] = sum(b["pli_premium"] for b in s["bos"])
            s["pli"]["non_procurers"] = sum(1 for b in s["bos"] if b["pli_policies"] == 0)
            s["rpli"]["offices_procured"] = rpli_off
            s["rpli"]["policies"] = sum(b["rpli_policies"] for b in s["bos"])
            s["rpli"]["premium"] = sum(b["rpli_premium"] for b in s["bos"])
            s["rpli"]["non_procurers"] = sum(1 for b in s["bos"] if b["rpli_policies"] == 0)

    subdiv_out = {div: {"subdivs": {sd: dict(s) for sd, s in subs.items()}} for div, subs in by_div.items()}

    # ── Search index for BO_LOOKUP ──
    search_idx = []
    for oid, r in bo_lookup.items():
        name_n = norm_name(r["name"])
        search_idx.append({
            "id": oid, "n": r["name"], "c": r["code"], "t": r["type"],
            "d": r["division"], "r": r["region"], "sd": r["sub_division"],
            "k": f"{name_n} {(r['code'] or '').lower()} {oid}".strip(),
        })

    (BASE / "_subdiv_data_june.json").write_text(json.dumps(subdiv_out, ensure_ascii=False), encoding="utf-8")
    (BASE / "_bo_lookup_june.json").write_text(
        json.dumps({"lookup": bo_lookup, "index": search_idx}, ensure_ascii=False), encoding="utf-8")

    total_offices = sum(s["offices"] for d in by_div.values() for s in d.values())
    print(f"\nSUBDIV_DATA: {len(by_div)} divisions, {sum(len(d) for d in by_div.values())} subdivisions, {total_offices} offices")
    print(f"BO_LOOKUP: {len(bo_lookup):,} offices")

    # sanity totals
    tot_posb_net = sum(r["posb"]["net"] for r in bo_lookup.values())
    tot_pli_pol = sum(r["pli"]["policies"] for r in bo_lookup.values())
    tot_rpli_pol = sum(r["rpli"]["policies"] for r in bo_lookup.values())
    print(f"POSB net (Apr+May+Jun): {tot_posb_net:+,}")
    print(f"PLI policies (Apr+May+Jun): {tot_pli_pol:,}  (expect 15,060)")
    print(f"RPLI policies (Apr+May+Jun): {tot_rpli_pol:,}  (expect 39,401)")


if __name__ == "__main__":
    main()

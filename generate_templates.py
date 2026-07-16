#!/usr/bin/env python3
"""generate_templates.py — builds /templates/<SECTION>.xlsx|csv from
pipeline_config.yaml. Templates and pipeline/validate.py both read the same
config, so a template can never drift out of sync with what CI validates.

Run whenever pipeline_config.yaml's header definitions change:
    python generate_templates.py
"""
import csv
from pathlib import Path
import openpyxl

from pipeline.config import load_config

BASE = Path(__file__).parent
TEMPLATES_DIR = BASE / "templates"

NOTE = ("Fill in your data below the header row shown here. Columns not "
        "listed are ignored by the dashboard pipeline — you don't need to "
        "remove them if your source export includes extra columns, but do "
        "not rename or reorder the columns listed below.")


def _write_xlsx_sheet(ws, sheet_cfg: dict):
    header_row = 1 if sheet_cfg["header_row"] in ("dynamic", 1) else sheet_cfg["header_row"]
    specs = sheet_cfg["required_headers"]
    if any(spec.get("column") is not None for spec in specs):
        # Column-pinned layout (e.g. ECR sheet, where position matters).
        for spec in specs:
            col = spec.get("column")
            if col is not None:
                ws.cell(row=header_row, column=col + 1, value=spec["name"])
    else:
        # No pinned columns (e.g. SWAP, POSB, PLI/RPLI sheets) — lay
        # headers out left to right in the order given.
        for i, spec in enumerate(specs):
            ws.cell(row=header_row, column=i + 1, value=spec["name"])


def generate_posb_silent_template(section_cfg: dict):
    """posb_silent's meta/data shape doesn't fit the header-row model every
    other section uses (see pipeline/sections/posb_silent.py) — the meta
    sheet here is left blank apart from labels, since schema_version,
    row_count, and data_sha256_16 all have to be computed from the
    operator's own finished data sheet, not filled in by hand."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    meta_ws = wb.create_sheet("meta")
    meta_ws.append(["key", "value"])
    for key, note in [
        ("schema_version", "1"),
        ("as_of_month", "YYYY-MM, e.g. 2026-06"),
        ("fy_start", "YYYY-MM, e.g. 2026-04"),
        ("months_covered", "comma-separated, e.g. 2026-04,2026-05,2026-06"),
        ("row_count", "must equal the data sheet's actual row count"),
        ("data_sha256_16", "first 16 hex chars of sha256 of the data sheet as CSV (headers, no index)"),
        ("generated", "YYYY-MM-DD HH:MM"),
    ]:
        meta_ws.append([key, note])

    data_ws = wb.create_sheet("data")
    data_ws.append(section_cfg["required_data_columns"])

    info = wb.create_sheet("README", 0)
    info["A1"] = f"{section_cfg['label']} — upload template"
    info["A2"] = (
        "This file's meta sheet must be generated programmatically, not filled in by hand — "
        "row_count and data_sha256_16 have to match the data sheet exactly or the upload is "
        "rejected. Fill in the data sheet, then run your generator script to (re)write the meta "
        "sheet from it before uploading."
    )
    out = TEMPLATES_DIR / "posb_silent.xlsx"
    wb.save(out)
    return out


def generate_xlsx_template(name: str, section_cfg: dict):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, sheet_cfg in section_cfg["sheets"].items():
        ws = wb.create_sheet(sheet_name)
        _write_xlsx_sheet(ws, sheet_cfg)
    info = wb.create_sheet("README", 0)
    info["A1"] = f"{section_cfg['label']} — upload template"
    info["A2"] = NOTE
    out = TEMPLATES_DIR / f"{name}.xlsx"
    wb.save(out)
    return out


def generate_csv_template(name: str, section_cfg: dict):
    out = TEMPLATES_DIR / f"{name}.csv"
    with open(out, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(section_cfg["csv_headers"]["required"])
    return out


def main():
    cfg = load_config()
    TEMPLATES_DIR.mkdir(exist_ok=True)
    for name, section_cfg in cfg["sections"].items():
        if name == "posb_silent":
            out = generate_posb_silent_template(section_cfg)
            print(f"  Wrote {out.relative_to(BASE)}")
            continue
        file_type = section_cfg.get("file_type", "xlsx")
        if file_type == "xlsx":
            out = generate_xlsx_template(name, section_cfg)
        elif file_type == "csv":
            out = generate_csv_template(name, section_cfg)
        else:
            print(f"  SKIP {name}: unknown file_type {file_type!r}")
            continue
        print(f"  Wrote {out.relative_to(BASE)}")


if __name__ == "__main__":
    main()

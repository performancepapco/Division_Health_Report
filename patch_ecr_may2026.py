#!/usr/bin/env python3
"""
patch_ecr_may2026.py
────────────────────
Reads uploads/may_2026/ECR_Calculation_May2026.xlsx and injects cumulative
(Apr + May) revenue, expenditure, and SWAP breakdown values into
DATA_BY_MONTH['May-26'] in index.html so the dashboard recomputes ECR
correctly for the May-26 snapshot.

Run:  python patch_ecr_may2026.py
"""

import json, re
from pathlib import Path
import openpyxl

BASE = Path(__file__).parent
HTML = BASE / "index.html"
XLSX = BASE / "uploads" / "may_2026" / "ECR_Calculation_May2026.xlsx"
CR = 10_000_000  # 1 Cr

# Divisions present in DATA_BY_MONTH (must match exactly)
DIVISIONS = {
    "Amalapuram", "Anakapalle", "Anantapur", "Bhimavaram", "Chittoor",
    "Cuddapah", "Eluru", "Gudivada", "Gudur", "Guntur", "Hindupur",
    "Kakinada", "Kurnool", "Machilipatnam", "Markapur", "Nandyal",
    "Narasaraopet", "Nellore", "Parvathipuram", "Prakasam", "Proddatur",
    "RMS AG", "RMS TP", "RMS V", "RMS Y", "Rajahmundry", "Srikakulam",
    "Tadepalligudem", "Tenali", "Tirupati", "Vijayawada", "Visakhapatnam",
    "Vizianagaram",
}

# ECR sheet column indices (0-based)
COL_DIV   = 0
COL_POSB  = 8   # Total rev on sbcc up to May-2026
COL_PLIR  = 24  # Total PLI/RPLI Rev.
COL_PARC  = 25  # Parcel
COL_MO    = 26  # MO
COL_IRGB  = 27  # IRGB
COL_CCS   = 28  # CCS
COL_TOT_R = 31  # Total revenue
COL_TOT_E = 34  # Exp. For ECR calculation
COL_PEN   = 36  # 320107 (pension component)

# SWAP sheet columns
S_DIV   = 0
S_SAL   = 1
S_WAGES = 2
S_ALLOW = 3
S_PEN   = 4
S_OTHER = 5
S_PLAN  = 6
S_TOTAL = 7


def to_cr(v):
    return round((v or 0) / CR, 7)


def read_ecr():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ecr_ws = wb["ECR"]
    swap_ws = wb["SWAP"]

    out = {}

    for row in ecr_ws.iter_rows(min_row=2, values_only=True):
        name = row[COL_DIV]
        if not name or name not in DIVISIONS:
            continue
        out[name] = {
            "actuals": {
                "Parcel":          to_cr(row[COL_PARC]),
                "Mail Operations": to_cr(row[COL_MO]),
                "IR & GB":         to_cr(row[COL_IRGB]),
                "CCS":             to_cr(row[COL_CCS]),
                "POSB":            to_cr(row[COL_POSB]),
                "PLI/RPLI":        to_cr(row[COL_PLIR]),
                "Total":           to_cr(row[COL_TOT_R]),
            },
            "total_exp": to_cr(row[COL_TOT_E]),
        }

    for row in swap_ws.iter_rows(min_row=2, values_only=True):
        name = row[S_DIV]
        if not name or name not in DIVISIONS:
            continue
        if name not in out:
            continue
        out[name]["swap"] = {
            "Salaries":   to_cr(row[S_SAL]),
            "Wages":      to_cr(row[S_WAGES]),
            "Pension":    to_cr(row[S_PEN]),
            "Allowances": to_cr(row[S_ALLOW]),
            "Plan":       to_cr(row[S_PLAN]),
            "Other":      to_cr(row[S_OTHER]),
            "Total":      to_cr(row[S_TOTAL]),
        }

    return out


def build_injection(ecr_data):
    js = (
        "\n\n// === ECR May-26 cumulative financials (auto-generated) ===\n"
        "(function(){\n"
        "  var ECR_MAY = " + json.dumps(ecr_data, ensure_ascii=False) + ";\n"
        "  if (!DATA_BY_MONTH || !DATA_BY_MONTH['May-26']) return;\n"
        "  Object.keys(ECR_MAY).forEach(function(k) {\n"
        "    var d = DATA_BY_MONTH['May-26'][k];\n"
        "    if (!d) return;\n"
        "    var e = ECR_MAY[k];\n"
        "    d.actuals   = e.actuals;\n"
        "    d.total_exp = e.total_exp;\n"
        "    d.swap      = e.swap;\n"
        "  });\n"
        "})();\n"
        "// === END ECR May-26 ===\n"
    )
    return js


MARKER_BEGIN = "// === ECR May-26 cumulative financials (auto-generated) ==="
MARKER_END   = "// === END ECR May-26 ==="


def patch_html(injection):
    html = HTML.read_text(encoding="utf-8")

    # Idempotent: replace any prior injection block
    if MARKER_BEGIN in html:
        pattern = re.compile(
            re.escape(MARKER_BEGIN) + r".*?" + re.escape(MARKER_END) + r"\n",
            re.DOTALL,
        )
        new_html, n = pattern.subn(injection.lstrip("\n"), html, count=1)
        print(f"  Replaced existing ECR injection ({n} occurrence)")
        html = new_html
    else:
        # Insert after the May-26 booking-aggregates IIFE inserted by patch_may2026.py.
        # That block ends with: })(); just before the next "// ===" or "</script>".
        # We anchor on the existing May-26 booking IIFE close.
        anchor = "DATA_BY_MONTH['May-26'] = JSON.parse"
        idx = html.find(anchor)
        if idx == -1:
            print("  ERROR: Could not find May-26 init block. Run patch_may2026.py first.")
            return False
        # Find the closing })(); of that IIFE (first one after the anchor)
        close = html.find("})();", idx)
        if close == -1:
            print("  ERROR: Could not find closing of May-26 IIFE.")
            return False
        insert_at = close + len("})();")
        html = html[:insert_at] + injection + html[insert_at:]
        print("  Inserted ECR injection after May-26 booking block")

    HTML.write_text(html, encoding="utf-8")
    return True


if __name__ == "__main__":
    print("=== patch_ecr_may2026.py ===")
    print(f"Reading {XLSX.name}...")
    ecr_data = read_ecr()
    print(f"  Loaded ECR cumulative values for {len(ecr_data)} divisions")

    missing = DIVISIONS - set(ecr_data.keys())
    if missing:
        print(f"  WARNING: missing in Excel: {sorted(missing)}")

    print("Building JS injection...")
    inj = build_injection(ecr_data)

    print("Patching index.html...")
    if patch_html(inj):
        print("\n=== Done ===")
        print("Open index.html, switch month-to=May-2026, and check ECR values.")

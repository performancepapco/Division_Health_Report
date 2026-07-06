#!/usr/bin/env python3
"""
build_pli_rpli_june2026.py
───────────────────────────
Reads standalone-month PLI/RPLI Distribution files (April, May in
uploads/june_2026/Insurance/, June in uploads/june_2026/) and builds:
  - division-level policies/premium per month (for PLI_POLICIES/RPLI_POLICIES tables)
  - office-level Detail per month (for Sub-Div Drilldown / BO Lookup cumulative build)

Writes:
  _pli_rpli_by_month.json   division-level, keyed [segment][month][division]
  _pli_rpli_offices.json    office-level, keyed [segment][month] -> list of records

Run:  python build_pli_rpli_june2026.py
"""
import json, re
from pathlib import Path
from collections import defaultdict
import openpyxl

BASE = Path(__file__).parent
INS = BASE / "uploads" / "june_2026" / "Insurance"
JUN = BASE / "uploads" / "june_2026"

FILES = {
    "pli": {
        "Apr-26": (INS / "PLI_April2026_Distribution.xlsx", "PLI Division Distribution", "PLI Detail"),
        "May-26": (INS / "PLI_May2026_Distribution.xlsx", "PLI Division Distribution", "PLI Detail"),
        "Jun-26": (JUN / "PLI_June2026_Distribution.xlsx", "PLI Division Distribution", "PLI Detail"),
    },
    "rpli": {
        "Apr-26": (INS / "RPLI_April2026_Distribution.xlsx", "RPLI Division Distribution", "RPLI Detail"),
        "May-26": (INS / "RPLI_May2026_Distribution.xlsx", "RPLI Division Distribution", "RPLI Detail"),
        "Jun-26": (JUN / "RPLI_June2026_Distribution.xlsx", "RPLI Division Distribution", "RPLI Detail"),
    },
}


def short_div(s):
    return re.sub(r"\s+Division\s*$", "", str(s)).strip() if s else ""


def read_division_level(path, sheet):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    # header row is the one starting with 'Sl.'
    hdr_idx = next(i for i, r in enumerate(rows) if r[0] == "Sl.")
    out = {}
    circle = None
    for r in rows[hdr_idx + 1:]:
        if r[1] is None:
            continue
        div_raw = r[1]
        total_off, off_proc = r[3] or 0, r[4] or 0
        policies, premium = r[12] or 0, r[13] or 0
        region = r[2] or ""
        if str(div_raw).strip() == "CIRCLE TOTAL":
            circle = {
                "total_offices": int(total_off), "offices_procured": int(off_proc),
                "policies": int(policies), "premium": float(premium),
            }
            continue
        div = short_div(div_raw)
        procurement_rate = round(off_proc / total_off * 100, 1) if total_off else 0
        avg_premium = round(premium / policies) if policies else 0
        out[div] = {
            "total_offices": int(total_off), "offices_procured": int(off_proc),
            "policies": int(policies), "premium": float(premium),
            "procurement_rate": procurement_rate, "avg_premium": avg_premium,
            "region": str(region).replace(" Region", "").strip(),
        }
    return out, circle


HEADER_RE = re.compile(r"^(.+?)\s+Division\s")


def read_office_level(path, sheet):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    cur_div = None
    records = []
    unmatched_div_rows = 0
    for row in ws.iter_rows(values_only=True):
        a = row[0]
        if a is None:
            continue
        if isinstance(a, str):
            m = HEADER_RE.match(a)
            if m and "band" in a.lower():
                cur_div = short_div(m.group(1).strip())
                continue
            if a == "Sl.":
                continue
        if not isinstance(row[0], int):
            continue
        if cur_div is None:
            unmatched_div_rows += 1
            continue
        code, name, otype, pols, prem = row[1], row[2], row[3], row[4], row[5]
        if not name:
            continue
        records.append({
            "division": cur_div, "code": code, "name": name, "type": otype or "",
            "policies": int(pols or 0), "premium": float(prem or 0),
        })
    return records, unmatched_div_rows


CUM_FILES = {
    "pli":  (JUN / "PLI_AprToJune2026_Cumulative_Distribution.xlsx", "PLI Division Distribution"),
    "rpli": (JUN / "RPLI_AprToJune2026_Cumulative_Distribution.xlsx", "RPLI Division Distribution"),
}


def main():
    by_month = {"pli": {}, "rpli": {}}
    offices = {"pli": {}, "rpli": {}}
    cumulative = {"pli": {}, "rpli": {}}

    for segment, (path, sheet) in CUM_FILES.items():
        print(f"[{segment}] Apr-to-Jun cumulative: {path.name}")
        div_data, circle = read_division_level(path, sheet)
        div_data["__CIRCLE__"] = circle
        cumulative[segment] = div_data
        print(f"  circle policies={circle['policies']:,} premium={circle['premium']:,.0f}")

    for segment, months in FILES.items():
        for month, (path, div_sheet, detail_sheet) in months.items():
            print(f"[{segment}] {month}: {path.name}")
            div_data, circle = read_division_level(path, div_sheet)
            div_data["__CIRCLE__"] = circle
            by_month[segment][month] = div_data
            print(f"  {len(div_data)-1} divisions, circle policies={circle['policies']:,} premium={circle['premium']:,.0f}")

            recs, unmatched = read_office_level(path, detail_sheet)
            offices[segment][month] = recs
            print(f"  {len(recs)} office rows ({unmatched} rows skipped before first division header)")

    (BASE / "_pli_rpli_by_month.json").write_text(
        json.dumps(by_month, ensure_ascii=False), encoding="utf-8")
    (BASE / "_pli_rpli_offices.json").write_text(
        json.dumps(offices, ensure_ascii=False), encoding="utf-8")
    (BASE / "_pli_rpli_cumulative.json").write_text(
        json.dumps(cumulative, ensure_ascii=False), encoding="utf-8")
    print("\nWrote _pli_rpli_by_month.json, _pli_rpli_offices.json, _pli_rpli_cumulative.json")

    # Reconciliation check: Apr + May + Jun standalone vs Apr-to-Jun cumulative file
    for seg in ("pli", "rpli"):
        tot_pol = sum(by_month[seg][m]["__CIRCLE__"]["policies"] for m in ("Apr-26", "May-26", "Jun-26"))
        tot_prem = sum(by_month[seg][m]["__CIRCLE__"]["premium"] for m in ("Apr-26", "May-26", "Jun-26"))
        print(f"{seg.upper()} Apr+May+Jun -> policies={tot_pol:,} premium={tot_prem:,.0f}")


if __name__ == "__main__":
    main()

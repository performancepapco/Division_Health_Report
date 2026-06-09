#!/usr/bin/env python3
"""
patch_subdiv_may2026.py
───────────────────────
Builds sub-division × office-level data for POSB, PLI, RPLI from the
May-2026 uploads, then injects a SUBDIV_DATA JS constant into index.html
and renders:
  - Inline sub-division summary in POSB / PLI / RPLI tabs (per selected division)
  - A new top-level "Sub-Div Drilldown" tab with division picker
  - "BOs that did not procure" listed for each segment (POSB/PLI/RPLI)

Run:  python patch_subdiv_may2026.py
"""

import re, json
from pathlib import Path
from collections import defaultdict
import openpyxl

BASE = Path(__file__).parent
HTML = BASE / "index.html"
UP   = BASE / "uploads" / "may_2026"

POSB_XLSX = UP / "POSB Report April and May 2026.xlsx"
PLI_XLSX  = UP / "PLI_MAY2026_Drilldown.xlsx"
RPLI_XLSX = UP / "RPLI_MAY2026_Drilldown.xlsx"


def norm_name(s):
    """Normalize an office name for cross-source joins."""
    if not s:
        return ""
    s = str(s).strip().lower()
    s = re.sub(r"\s+", " ", s)
    s = s.replace(".", "").replace(",", "")
    s = s.replace(" b.o", " bo").replace(" s.o", " so").replace(" h.o", " ho")
    s = re.sub(r"\bb\.?o\b", "bo", s)
    s = re.sub(r"\bs\.?o\b", "so", s)
    s = re.sub(r"\bh\.?o\b", "ho", s)
    return s.strip()


def short_div(div_name):
    """'Anantapur Division' → 'Anantapur'; 'Cuddapah' → 'Cuddapah'."""
    return re.sub(r"\s+Division\s*$", "", str(div_name)).strip()


def read_posb():
    """Returns: division → sub_division → {summary, bos[]}."""
    wb = openpyxl.load_workbook(POSB_XLSX, data_only=True)

    # Sheet 2: per-BO totals with sub-division
    ws = wb["2. BO Totals"]
    bo_map = {}   # (div, name_norm) → bo record
    by_div = defaultdict(lambda: defaultdict(lambda: {
        "region": "", "offices": 0,
        "posb": {"opened": 0, "closed": 0, "net": 0},
        "pli":  {"offices_procured": 0, "policies": 0, "premium": 0, "non_procurers": 0},
        "rpli": {"offices_procured": 0, "policies": 0, "premium": 0, "non_procurers": 0},
        "bos": [],
    }))
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[1]:
            continue
        oid, otype, code, _, name, sub_div, div, opened, closed, net, region = (
            row[1], row[2], row[3], row[4], row[5], row[6], row[7],
            row[8] or 0, row[9] or 0, row[10] or 0, row[11] or ""
        )
        div_s = short_div(div)
        sd = sub_div or "(Unmapped)"
        rec = {
            "id": oid, "code": code, "name": name, "type": otype,
            "posb_opened": int(opened), "posb_closed": int(closed), "posb_net": int(net),
            "pli_policies": 0, "pli_premium": 0,
            "rpli_policies": 0, "rpli_premium": 0,
        }
        bo_map[(div_s, norm_name(name))] = rec
        s = by_div[div_s][sd]
        s["region"] = region.replace(" Region", "") if region else ""
        s["offices"] += 1
        s["posb"]["opened"] += rec["posb_opened"]
        s["posb"]["closed"] += rec["posb_closed"]
        s["posb"]["net"]    += rec["posb_net"]
        s["bos"].append(rec)
    return by_div, bo_map


def read_drilldown(xlsx_path, segment, bo_map, by_div):
    """Parse PLI/RPLI drilldown Detail sheet → update bo_map / aggregates."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Detail"]
    cur_div = None
    pol_key  = f"{segment}_policies"
    prem_key = f"{segment}_premium"

    header_re = re.compile(r"^(.+?)\s+Division\s+—\s+policy band", re.IGNORECASE)
    unmatched = 0

    for row in ws.iter_rows(min_row=1, values_only=True):
        a = row[0]
        if not a:
            continue
        m = header_re.match(str(a))
        if m:
            cur_div = short_div(m.group(1).strip())
            continue
        if str(a) == "Sl.":
            continue
        if not isinstance(row[0], int):
            continue
        # data row: (Sl, OfficeCode, OfficeName, OfficeType, Policies, Premium)
        if cur_div is None:
            continue
        name = row[2]
        pols = int(row[4] or 0)
        prem = float(row[5] or 0)
        if not name:
            continue
        key = (cur_div, norm_name(name))
        rec = bo_map.get(key)
        if rec is None:
            unmatched += 1
            continue
        rec[pol_key]  = pols
        rec[prem_key] = prem
    return unmatched


def aggregate_segments(by_div):
    """Recompute sub-division PLI/RPLI totals from updated BO records."""
    for div, subs in by_div.items():
        for sd, s in subs.items():
            pli_off  = sum(1 for b in s["bos"] if b["pli_policies"] > 0)
            rpli_off = sum(1 for b in s["bos"] if b["rpli_policies"] > 0)
            s["pli"]["offices_procured"] = pli_off
            s["pli"]["policies"]         = sum(b["pli_policies"] for b in s["bos"])
            s["pli"]["premium"]          = sum(b["pli_premium"]  for b in s["bos"])
            s["pli"]["non_procurers"]    = sum(1 for b in s["bos"] if b["pli_policies"] == 0)
            s["rpli"]["offices_procured"]= rpli_off
            s["rpli"]["policies"]        = sum(b["rpli_policies"] for b in s["bos"])
            s["rpli"]["premium"]         = sum(b["rpli_premium"]  for b in s["bos"])
            s["rpli"]["non_procurers"]   = sum(1 for b in s["bos"] if b["rpli_policies"] == 0)


def build_data():
    print("  Reading POSB...")
    by_div, bo_map = read_posb()
    total_bos = sum(s["offices"] for d in by_div.values() for s in d.values())
    print(f"    {len(by_div)} divisions, {sum(len(d) for d in by_div.values())} sub-divisions, {total_bos} offices")

    print("  Reading PLI drilldown...")
    pli_un = read_drilldown(PLI_XLSX, "pli", bo_map, by_div)
    print(f"    Unmatched PLI offices: {pli_un}")

    print("  Reading RPLI drilldown...")
    rpli_un = read_drilldown(RPLI_XLSX, "rpli", bo_map, by_div)
    print(f"    Unmatched RPLI offices: {rpli_un}")

    print("  Aggregating sub-division totals...")
    aggregate_segments(by_div)

    # Convert defaultdicts to plain dicts for json
    out = {}
    for div, subs in by_div.items():
        out[div] = {"subdivs": {sd: dict(s) for sd, s in subs.items()}}
    return out


# ─── HTML / CSS / JS injection ────────────────────────────────────────────────

MARKER_BEGIN_DATA = "// === SUBDIV_DATA (auto-generated) ==="
MARKER_END_DATA   = "// === END SUBDIV_DATA ==="

MARKER_BEGIN_CSS  = "/* === SUBDIV STYLES === */"
MARKER_END_CSS    = "/* === END SUBDIV STYLES === */"

MARKER_BEGIN_JS   = "// === SUBDIV RENDER (auto-generated) ==="
MARKER_END_JS     = "// === END SUBDIV RENDER ==="

MARKER_BEGIN_TAB  = "<!-- === SUBDIV TAB === -->"
MARKER_END_TAB    = "<!-- === END SUBDIV TAB === -->"

NEW_CSS = """
/* === SUBDIV STYLES === */
.subdiv-wrap { margin-top: 22px; }
.subdiv-head {
  display: flex; align-items: baseline; justify-content: space-between;
  margin-bottom: 10px; padding-bottom: 6px;
  border-bottom: 1px solid var(--rule);
}
.subdiv-head h3 {
  font-family: var(--font-mono); font-size: 13px; letter-spacing: 0.16em;
  text-transform: uppercase; color: var(--ink-soft); margin: 0;
}
.subdiv-head .subdiv-count {
  font-family: var(--font-mono); font-size: 12px; color: var(--ink-faint);
}
.subdiv-table { width: 100%; border-collapse: collapse; font-family: var(--font-sans); font-size: 13px; }
.subdiv-table th {
  font-family: var(--font-mono); font-size: 11px; letter-spacing: 0.12em;
  text-transform: uppercase; color: var(--ink-soft); font-weight: 500;
  text-align: right; padding: 6px 8px; border-bottom: 1px solid var(--rule);
  background: var(--paper-warm);
}
.subdiv-table th:first-child, .subdiv-table td:first-child { text-align: left; }
.subdiv-table td { padding: 6px 8px; border-bottom: 1px solid var(--rule); text-align: right; }
.subdiv-table tr:hover td { background: var(--paper-warm); cursor: pointer; }
.subdiv-table .num.pos { color: var(--green); font-weight: 600; }
.subdiv-table .num.neg { color: var(--red); font-weight: 600; }
.subdiv-table .num.zero { color: var(--ink-faint); }
.subdiv-table .badge {
  font-family: var(--font-mono); font-size: 10.5px; padding: 1px 5px;
  border-radius: 2px; background: var(--paper-warm); color: var(--ink-soft);
  letter-spacing: 0.04em; margin-left: 4px;
}
.subdiv-table .badge.warn { background: var(--amber-pale); color: oklch(0.45 0.13 70); }
.subdiv-table .badge.danger { background: var(--red-pale); color: oklch(0.45 0.13 25); }

.bo-list-wrap {
  background: var(--paper-warm); padding: 10px 14px;
  border-left: 3px solid var(--accent); margin: 4px 0 12px;
}
.bo-list-wrap h5 {
  font-family: var(--font-mono); font-size: 12px; letter-spacing: 0.12em;
  text-transform: uppercase; color: var(--ink-soft); margin: 0 0 6px;
}
.bo-list-table { width: 100%; border-collapse: collapse; font-size: 12.5px; font-family: var(--font-sans); }
.bo-list-table th, .bo-list-table td { padding: 4px 6px; border-bottom: 1px solid var(--rule); text-align: right; }
.bo-list-table th { font-family: var(--font-mono); font-size: 10.5px; letter-spacing: 0.10em; text-transform: uppercase; color: var(--ink-soft); font-weight: 500; }
.bo-list-table th:nth-child(-n+3), .bo-list-table td:nth-child(-n+3) { text-align: left; }
.bo-list-table tr.nil-row td { color: var(--ink-faint); }
.bo-list-table tr.nil-row td.flag { color: var(--red); font-weight: 600; }
.bo-filters {
  display: flex; gap: 8px; margin: 6px 0 10px; align-items: center;
  font-size: 12px; font-family: var(--font-mono);
}
.bo-filters label { color: var(--ink-soft); }
.bo-filters select {
  font-size: 12px; padding: 2px 4px;
  border: 1px solid var(--rule); border-radius: 3px;
  background: var(--surface-1); color: var(--ink);
}

.sd-tab-controls {
  display: flex; gap: 16px; align-items: center;
  margin-bottom: 16px; padding: 10px 14px;
  background: var(--paper-warm); border: 1px solid var(--rule);
}
.sd-tab-controls label { font-family: var(--font-mono); font-size: 12px; color: var(--ink-soft); letter-spacing: 0.06em; }
.sd-tab-controls select {
  font-size: 13px; padding: 4px 6px;
  border: 1px solid var(--rule); border-radius: 3px;
  background: var(--surface-1); color: var(--ink); font-family: var(--font-sans);
}
/* === END SUBDIV STYLES === */
"""

NEW_JS = r"""
// === SUBDIV RENDER (auto-generated) ===
(function(){
  function fmtN(n) {
    if (n === 0) return '0';
    var sign = n < 0 ? '-' : '';
    return sign + Math.abs(n).toLocaleString('en-IN');
  }
  function fmtPrem(n) {
    if (!n) return '0';
    if (n >= 1e7) return (n/1e7).toFixed(2) + ' Cr';
    if (n >= 1e5) return (n/1e5).toFixed(2) + ' L';
    return Math.round(n).toLocaleString('en-IN');
  }
  function cls(n) {
    if (n > 0) return 'pos';
    if (n < 0) return 'neg';
    return 'zero';
  }

  function rowsForSegment(division, segment) {
    var d = SUBDIV_DATA[division];
    if (!d) return '';
    var subs = d.subdivs;
    var keys = Object.keys(subs).sort();
    var html = '';
    keys.forEach(function(sd) {
      var s = subs[sd];
      var seg = s[segment];
      if (segment === 'posb') {
        html += '<tr data-div="'+division+'" data-sd="'+sd+'" data-seg="posb" onclick="toggleBOList(this)">'
              + '<td>' + sd + '</td>'
              + '<td>' + s.offices + '</td>'
              + '<td>' + fmtN(s.posb.opened) + '</td>'
              + '<td>' + fmtN(s.posb.closed) + '</td>'
              + '<td class="num '+cls(s.posb.net)+'">' + (s.posb.net>=0?'+':'') + fmtN(s.posb.net) + '</td>'
              + '</tr>';
      } else {
        var nonPr = seg.non_procurers || 0;
        var badge = nonPr > 0
          ? '<span class="badge '+(nonPr/s.offices>0.7?'danger':'warn')+'">'+nonPr+' nil</span>'
          : '';
        html += '<tr data-div="'+division+'" data-sd="'+sd+'" data-seg="'+segment+'" onclick="toggleBOList(this)">'
              + '<td>' + sd + ' ' + badge + '</td>'
              + '<td>' + s.offices + '</td>'
              + '<td>' + seg.offices_procured + '</td>'
              + '<td>' + fmtN(seg.policies) + '</td>'
              + '<td>' + fmtPrem(seg.premium) + '</td>'
              + '</tr>';
      }
    });
    return html;
  }

  window.renderSubdivInline = function(divName, segment, mountId) {
    var el = document.getElementById(mountId);
    if (!el) return;
    if (!divName || divName === '__CIRCLE__' || !SUBDIV_DATA[divName]) {
      el.innerHTML = '';
      return;
    }
    var heads = (segment === 'posb')
      ? '<tr><th>Sub-Division</th><th>Offices</th><th>Opened</th><th>Closed</th><th>Net Addition</th></tr>'
      : '<tr><th>Sub-Division</th><th>Offices</th><th>Procured</th><th>Policies</th><th>Premium</th></tr>';
    var rows = rowsForSegment(divName, segment);
    el.innerHTML =
      '<div class="subdiv-wrap">'
      + '<div class="subdiv-head"><h3>Sub-Division breakdown</h3>'
      + '<span class="subdiv-count">Click any row to see Branch Offices</span></div>'
      + '<table class="subdiv-table"><thead>' + heads + '</thead>'
      + '<tbody>' + rows + '</tbody></table></div>';
  };

  window.toggleBOList = function(tr) {
    var div = tr.dataset.div, sd = tr.dataset.sd, seg = tr.dataset.seg;
    var next = tr.nextElementSibling;
    if (next && next.classList.contains('bo-list-row')) {
      next.remove();
      return;
    }
    // close any other open row in same table
    var tbody = tr.parentElement;
    tbody.querySelectorAll('tr.bo-list-row').forEach(function(r){ r.remove(); });

    var s = SUBDIV_DATA[div].subdivs[sd];
    var bos = s.bos.slice();
    // sort: non-procurers first if seg is pli/rpli, else by net desc
    if (seg === 'posb') bos.sort(function(a,b){ return b.posb_net - a.posb_net; });
    else if (seg === 'pli') bos.sort(function(a,b){ return b.pli_policies - a.pli_policies; });
    else bos.sort(function(a,b){ return b.rpli_policies - a.rpli_policies; });

    var headRow, bodyRows = '';
    if (seg === 'posb') {
      headRow = '<tr><th>Office</th><th>Code</th><th>Type</th><th>Opened</th><th>Closed</th><th>Net</th></tr>';
      bos.forEach(function(b){
        var nil = (b.posb_opened === 0 && b.posb_closed === 0);
        bodyRows += '<tr class="'+(nil?'nil-row':'')+'">'
          + '<td>'+b.name+'</td><td>'+b.code+'</td><td>'+b.type+'</td>'
          + '<td>'+fmtN(b.posb_opened)+'</td><td>'+fmtN(b.posb_closed)+'</td>'
          + '<td class="'+(nil?'flag':'num '+cls(b.posb_net))+'">'+(nil?'nil':(b.posb_net>=0?'+':'')+fmtN(b.posb_net))+'</td>'
          + '</tr>';
      });
    } else {
      var polK = seg+'_policies', prK = seg+'_premium';
      headRow = '<tr><th>Office</th><th>Code</th><th>Type</th><th>Policies</th><th>Premium</th></tr>';
      bos.forEach(function(b){
        var nil = b[polK] === 0;
        bodyRows += '<tr class="'+(nil?'nil-row':'')+'">'
          + '<td>'+b.name+'</td><td>'+b.code+'</td><td>'+b.type+'</td>'
          + '<td class="'+(nil?'flag':'')+'">'+(nil?'nil':fmtN(b[polK]))+'</td>'
          + '<td>'+(nil?'—':fmtPrem(b[prK]))+'</td>'
          + '</tr>';
      });
    }
    var ncolspan = (seg === 'posb') ? 5 : 5;
    var newRow = document.createElement('tr');
    newRow.className = 'bo-list-row';
    var nonCount = (seg === 'posb')
      ? bos.filter(function(b){return b.posb_opened===0 && b.posb_closed===0;}).length
      : bos.filter(function(b){return b[seg+'_policies']===0;}).length;
    newRow.innerHTML = '<td colspan="'+ncolspan+'" style="padding:0; background:transparent;">'
      + '<div class="bo-list-wrap">'
      + '<h5>'+sd+' · Branch Offices ('+bos.length+' total · '+nonCount+' with nil '+(seg==='posb'?'activity':'procurement')+')</h5>'
      + '<table class="bo-list-table"><thead>'+headRow+'</thead><tbody>'+bodyRows+'</tbody></table>'
      + '</div></td>';
    tr.parentElement.insertBefore(newRow, tr.nextSibling);
  };

  // Sub-Div Drilldown tab
  window.renderSubdivTab = function() {
    var el = document.getElementById('subdiv-tab-content');
    if (!el) return;
    var pick = document.getElementById('sd-tab-div');
    var seg  = document.getElementById('sd-tab-seg');
    if (!pick || !seg) return;
    var div = pick.value;
    var s   = seg.value;
    if (!div || !SUBDIV_DATA[div]) { el.innerHTML = '<div class="placeholder">Select a division and segment above.</div>'; return; }
    var heads = (s === 'posb')
      ? '<tr><th>Sub-Division</th><th>Offices</th><th>Opened</th><th>Closed</th><th>Net Addition</th></tr>'
      : '<tr><th>Sub-Division</th><th>Offices</th><th>Procured</th><th>Policies</th><th>Premium</th></tr>';
    var rows = rowsForSegment(div, s);
    el.innerHTML = '<table class="subdiv-table"><thead>' + heads + '</thead><tbody>' + rows + '</tbody></table>';
  };

  window.initSubdivTab = function() {
    var pick = document.getElementById('sd-tab-div');
    if (!pick || pick.options.length > 0) return;
    var divs = Object.keys(SUBDIV_DATA).sort();
    pick.innerHTML = divs.map(function(d){ return '<option value="'+d+'">'+d+'</option>'; }).join('');
    renderSubdivTab();
  };
})();
// === END SUBDIV RENDER ===
"""

NEW_TAB_BTN = '<button class="tab" onclick="showTab(\'subdiv\', this); initSubdivTab();">Sub-Div Drilldown</button>'

NEW_TAB_HTML = """
<!-- === SUBDIV TAB === -->
<div id="tab-subdiv" class="tab-content">
  <div class="section-head">
    <h2>Sub-Division Drilldown — Apr–May 2026</h2>
    <span class="desc">Sub-division summaries by segment<br>Click any row to expand the Branch Office list (nil-procurement flagged)</span>
  </div>
  <div class="sd-tab-controls">
    <label for="sd-tab-div">Division:</label>
    <select id="sd-tab-div" onchange="renderSubdivTab()"></select>
    <label for="sd-tab-seg" style="margin-left:14px;">Segment:</label>
    <select id="sd-tab-seg" onchange="renderSubdivTab()">
      <option value="posb">POSB</option>
      <option value="pli">PLI</option>
      <option value="rpli">RPLI</option>
    </select>
  </div>
  <div id="subdiv-tab-content"></div>
</div>
<!-- === END SUBDIV TAB === -->
"""


def patch_html(subdiv_data):
    html = HTML.read_text(encoding="utf-8")
    print("  Read index.html ({:,} chars)".format(len(html)))

    # 1) Inject DATA constant
    data_block = (
        f"\n{MARKER_BEGIN_DATA}\n"
        f"const SUBDIV_DATA = {json.dumps(subdiv_data, ensure_ascii=False)};\n"
        f"{MARKER_END_DATA}\n"
    )
    if MARKER_BEGIN_DATA in html:
        pat = re.compile(re.escape(MARKER_BEGIN_DATA) + r".*?" + re.escape(MARKER_END_DATA) + r"\n",
                          re.DOTALL)
        html = pat.sub(data_block.lstrip("\n"), html, count=1)
        print("  Replaced SUBDIV_DATA")
    else:
        # Insert before the first existing PLI_POLICIES const (which is in script block 1)
        anchor = "const PLI_POLICIES"
        idx = html.find(anchor)
        if idx < 0:
            print("  ERROR: anchor PLI_POLICIES not found"); return False
        html = html[:idx] + data_block + html[idx:]
        print("  Inserted SUBDIV_DATA")

    # 2) Inject CSS
    if MARKER_BEGIN_CSS in html:
        pat = re.compile(re.escape(MARKER_BEGIN_CSS) + r".*?" + re.escape(MARKER_END_CSS),
                          re.DOTALL)
        html = pat.sub(NEW_CSS.strip(), html, count=1)
        print("  Replaced SUBDIV CSS")
    else:
        idx = html.rfind("</style>")
        html = html[:idx] + NEW_CSS + "\n" + html[idx:]
        print("  Inserted SUBDIV CSS")

    # 3) Inject JS render functions (into script block 1, before "function showTab")
    if MARKER_BEGIN_JS in html:
        pat = re.compile(re.escape(MARKER_BEGIN_JS) + r".*?" + re.escape(MARKER_END_JS) + r"\n",
                          re.DOTALL)
        html = pat.sub(NEW_JS.strip("\n") + "\n", html, count=1)
        print("  Replaced SUBDIV JS")
    else:
        # Place just before "function showTab(" in block 1
        m = re.search(r"\nfunction showTab\(", html)
        if not m:
            print("  ERROR: function showTab not found"); return False
        html = html[:m.start()] + "\n" + NEW_JS + "\n" + html[m.start():]
        print("  Inserted SUBDIV JS")

    # 4) Inject new tab button (after RPLI tab button)
    rpli_btn = "<button class=\"tab\" onclick=\"showTab('rpli', this)\">RPLI</button>"
    if "showTab('subdiv'" not in html:
        if rpli_btn in html:
            html = html.replace(rpli_btn, rpli_btn + "\n  " + NEW_TAB_BTN, 1)
            print("  Inserted Sub-Div tab button")
        else:
            print("  WARNING: RPLI tab button not found; sub-div button not inserted")

    # 5) Inject new tab HTML (after existing rpli tab div)
    if MARKER_BEGIN_TAB in html:
        pat = re.compile(re.escape(MARKER_BEGIN_TAB) + r".*?" + re.escape(MARKER_END_TAB),
                          re.DOTALL)
        html = pat.sub(NEW_TAB_HTML.strip(), html, count=1)
        print("  Replaced Sub-Div tab HTML")
    else:
        # Find balanced end of <div id="tab-rpli"> (it contains nested divs)
        i = html.find('<div id="tab-rpli"')
        if i < 0:
            print("  WARNING: tab-rpli div not found; tab HTML not inserted")
        else:
            depth, j = 0, i
            end_rpli = -1
            while j < len(html):
                no = html.find('<div', j)
                nc = html.find('</div>', j)
                if nc == -1:
                    break
                if no != -1 and no < nc:
                    depth += 1; j = no + 4
                else:
                    depth -= 1; j = nc + 6
                    if depth == 0:
                        end_rpli = j; break
            if end_rpli > 0:
                html = html[:end_rpli] + "\n\n" + NEW_TAB_HTML.strip() + "\n" + html[end_rpli:]
                print("  Inserted Sub-Div tab HTML")
            else:
                print("  WARNING: tab-rpli end not located; tab HTML not inserted")

    # 6) Hook inline sub-div renders into existing render functions
    # Add a div mount per existing tab content (idempotent)
    for tab, seg in [("posb", "posb"), ("pli", "pli"), ("rpli", "rpli")]:
        mount_id = f"subdiv-inline-{seg}"
        mount_div = f'  <div id="{mount_id}"></div>\n'
        if mount_id not in html:
            # Insert right before the closing </div> of the tab-<tab> div
            re_tab = re.compile(r'(<div id="tab-' + tab + r'"[^>]*>.*?)(\n</div>)', re.DOTALL)
            m = re_tab.search(html)
            if m:
                html = html[:m.start(2)] + "\n" + mount_div + html[m.start(2):]
                print(f"  Inserted inline mount for {tab}")

    # 7) Patch render functions to call renderSubdivInline at end
    # Find renderPOSBTab / renderPLITab / renderRPLITab and append the hook
    for fn, seg in [("renderPOSBTab", "posb"), ("renderPLITab", "pli"), ("renderRPLITab", "rpli")]:
        hook = f"  if (typeof renderSubdivInline === 'function') renderSubdivInline(divName, '{seg}', 'subdiv-inline-{seg}');\n"
        if hook in html:
            continue
        # Find "function <fn>(...) {" and append hook before the matching closing }
        pat = re.compile(r"(function " + re.escape(fn) + r"\([^)]*\)\s*\{)")
        m = pat.search(html)
        if not m:
            print(f"  WARNING: {fn} not found"); continue
        # Find matching closing brace at top level
        i = m.end()
        depth = 1
        while i < len(html) and depth > 0:
            c = html[i]
            if c == '{': depth += 1
            elif c == '}': depth -= 1
            i += 1
        # i is now just past the closing brace
        # Insert hook just before the closing brace
        insert_at = i - 1
        html = html[:insert_at] + hook + html[insert_at:]
        print(f"  Hooked subdiv into {fn}")

    HTML.write_text(html, encoding="utf-8")
    print("  Wrote index.html ({:,} chars)".format(len(html)))
    return True


if __name__ == "__main__":
    print("=== patch_subdiv_may2026.py ===")
    data = build_data()
    # Sanity print
    sample_div = next(iter(data))
    sample_sd = next(iter(data[sample_div]["subdivs"]))
    sd_obj = data[sample_div]["subdivs"][sample_sd]
    print(f"\nSample [{sample_div}/{sample_sd}]: "
          f"offices={sd_obj['offices']} posb_net={sd_obj['posb']['net']} "
          f"pli_proc={sd_obj['pli']['offices_procured']} rpli_proc={sd_obj['rpli']['offices_procured']}")

    print("\nPatching index.html...")
    if patch_html(data):
        print("\n=== Done ===")

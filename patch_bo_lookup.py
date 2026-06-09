#!/usr/bin/env python3
"""
patch_bo_lookup.py
──────────────────
Builds a per-office BO_LOOKUP dataset combining:
  - POSB activity (opened/closed/net) from POSB Report Apr+May
  - PLI / RPLI policies & premium from drilldown sheets
  - Booking activity per product from Apr + May booking CSVs (cumulative)
  - Pincode where available (POSB Sheet 9 Zero-Activity)
and injects a new "BO Lookup" tab into index.html with search + filters.

Run:  python patch_bo_lookup.py
"""

import re, json, csv
from pathlib import Path
from collections import defaultdict
import openpyxl

BASE = Path(__file__).parent
HTML = BASE / "index.html"
UP   = BASE / "uploads"
UPM  = UP / "may_2026"

POSB_XLSX = UPM / "POSB Report April and May 2026.xlsx"
PLI_XLSX  = UPM / "PLI_MAY2026_Drilldown.xlsx"
RPLI_XLSX = UPM / "RPLI_MAY2026_Drilldown.xlsx"
MAY_BK_CSV = UPM / "Booking_Productwise_Report (2).csv"
APR_BK_CSV = UP  / "booking_productwise.csv"


def norm_name(s):
    if not s: return ""
    s = str(s).strip().lower()
    s = re.sub(r"\s+", " ", s)
    s = s.replace(".", "").replace(",", "")
    s = re.sub(r"\bb\.?o\b", "bo", s)
    s = re.sub(r"\bs\.?o\b", "so", s)
    s = re.sub(r"\bh\.?o\b", "ho", s)
    return s.strip()


def short_div(s):
    return re.sub(r"\s+Division\s*$", "", str(s)).strip() if s else ""


def short_region(s):
    return str(s).replace(" Region", "").strip() if s else ""


# ─── 1. POSB master + activity ────────────────────────────────────────────────
def read_posb():
    print("  Reading POSB Sheet 2 (BO Totals)...")
    wb = openpyxl.load_workbook(POSB_XLSX, data_only=True)
    ws = wb["2. BO Totals"]
    lookup = {}
    name_map = {}   # (div, name_norm) → office_id
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[1]:
            continue
        oid = str(row[1])
        rec = {
            "id":   oid,
            "code": row[3] or "",
            "name": row[5] or "",
            "type": row[2] or "",
            "sub_division": row[6] or "",
            "division":     short_div(row[7]),
            "region":       short_region(row[11]),
            "pincode":      None,
            "posb": {"opened": int(row[8] or 0), "closed": int(row[9] or 0), "net": int(row[10] or 0)},
            "pli":  {"policies": 0, "premium": 0},
            "rpli": {"policies": 0, "premium": 0},
            "booking": {"total_articles": 0, "total_amount": 0, "products": []},
        }
        lookup[oid] = rec
        name_map[(rec["division"], norm_name(rec["name"]))] = oid

    print(f"    {len(lookup)} offices indexed from POSB report")

    # Pincode from Sheet 9
    try:
        ws9 = wb["9. Zero-Activity Offices"]
        n_pin = 0
        for row in ws9.iter_rows(min_row=2, values_only=True):
            oid = row[1]
            if oid is None: continue
            oid = str(oid)
            if oid in lookup and row[8]:
                lookup[oid]["pincode"] = str(row[8])
                n_pin += 1
        print(f"    Pincodes filled for {n_pin} zero-activity offices")
    except KeyError:
        print("    Sheet 9 not found, skipping pincodes")

    return lookup, name_map


# ─── 2. PLI / RPLI from drilldown ─────────────────────────────────────────────
def merge_drilldown(xlsx_path, segment, lookup, name_map):
    print(f"  Reading {xlsx_path.name}...")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Detail"]
    cur_div = None
    pol_k  = f"{segment}"
    header_re = re.compile(r"^(.+?)\s+Division\s+—\s+policy band", re.IGNORECASE)
    matched = unmatched = 0
    for row in ws.iter_rows(values_only=True):
        a = row[0]
        if not a:
            continue
        m = header_re.match(str(a))
        if m:
            cur_div = short_div(m.group(1).strip())
            continue
        if not isinstance(row[0], int):
            continue
        if cur_div is None: continue
        name = row[2]
        pols = int(row[4] or 0)
        prem = float(row[5] or 0)
        oid = name_map.get((cur_div, norm_name(name)))
        if oid is None:
            unmatched += 1; continue
        lookup[oid][pol_k]["policies"] = pols
        lookup[oid][pol_k]["premium"]  = prem
        matched += 1
    print(f"    {segment.upper()} matched: {matched}, unmatched: {unmatched}")


# ─── 3. Booking per office per product (Apr + May cumulative) ────────────────
def merge_booking(lookup):
    print("  Aggregating booking CSVs (Apr + May)...")
    bo_prod = defaultdict(lambda: defaultdict(lambda: {"articles": 0, "amount": 0.0}))

    for csv_path, label in [(APR_BK_CSV, "Apr"), (MAY_BK_CSV, "May")]:
        if not csv_path.exists():
            print(f"    WARN: {csv_path.name} not found, skipping {label}")
            continue
        n = 0
        with open(csv_path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                oid = (row.get("office-id") or "").strip()
                if not oid or oid == "-":
                    continue
                prod = (row.get("product-name") or "").strip()
                if not prod:
                    continue
                arts = int(row.get("article-count") or 0)
                amt  = float(row.get("total_amount") or 0)
                bo_prod[oid][prod]["articles"] += arts
                bo_prod[oid][prod]["amount"]   += amt
                n += 1
        print(f"    {label}: {n:,} rows aggregated")

    # Fold into lookup
    n_off = 0
    for oid, prod_map in bo_prod.items():
        if oid not in lookup:
            continue
        n_off += 1
        prods = []
        tot_a = 0; tot_amt = 0.0
        for pname, v in sorted(prod_map.items(), key=lambda kv: -kv[1]["amount"]):
            if v["articles"] == 0 and v["amount"] == 0:
                continue
            prods.append({"n": pname, "a": v["articles"], "v": round(v["amount"], 2)})
            tot_a += v["articles"]; tot_amt += v["amount"]
        lookup[oid]["booking"] = {
            "total_articles": tot_a,
            "total_amount":   round(tot_amt, 2),
            "products":       prods,
        }
    print(f"    Booking activity for {n_off} offices")


def build_search_index(lookup):
    """Compact search index — list of arrays for fast JS-side filtering."""
    idx = []
    for oid, r in lookup.items():
        name_n = norm_name(r["name"])
        idx.append({
            "id": oid,
            "n":  r["name"],
            "c":  r["code"],
            "t":  r["type"],
            "d":  r["division"],
            "r":  r["region"],
            "sd": r["sub_division"],
            "k":  f"{name_n} {(r['code'] or '').lower()} {oid}".strip(),
        })
    return idx


# ─── 4. HTML / CSS / JS injection ─────────────────────────────────────────────
MARK = {
    "data_b": "// === BO_LOOKUP DATA (auto-generated) ===",
    "data_e": "// === END BO_LOOKUP DATA ===",
    "css_b":  "/* === BO LOOKUP STYLES === */",
    "css_e":  "/* === END BO LOOKUP STYLES === */",
    "js_b":   "// === BO LOOKUP RENDER (auto-generated) ===",
    "js_e":   "// === END BO LOOKUP RENDER ===",
    "tab_b":  "<!-- === BO LOOKUP TAB === -->",
    "tab_e":  "<!-- === END BO LOOKUP TAB === -->",
}

NEW_CSS = """
/* === BO LOOKUP STYLES === */
.bolu-search-bar {
  display: grid;
  grid-template-columns: 1.6fr 1fr 1fr auto;
  gap: 10px;
  padding: 14px 16px;
  background: var(--paper-warm);
  border: 1px solid var(--rule);
  margin-bottom: 14px;
  align-items: end;
}
.bolu-field { display: flex; flex-direction: column; gap: 4px; }
.bolu-field label {
  font-family: var(--font-mono); font-size: 11px;
  letter-spacing: 0.12em; text-transform: uppercase; color: var(--ink-soft);
}
.bolu-field input, .bolu-field select {
  font-size: 13.5px; padding: 6px 8px;
  border: 1px solid var(--rule); border-radius: 3px;
  background: var(--surface-1); color: var(--ink);
  font-family: var(--font-sans);
}
.bolu-clear {
  font-family: var(--font-mono); font-size: 12px;
  padding: 7px 14px; height: fit-content;
  background: var(--paper); color: var(--ink-soft);
  border: 1px solid var(--rule); border-radius: 3px; cursor: pointer;
  letter-spacing: 0.08em; text-transform: uppercase;
}
.bolu-clear:hover { background: var(--ink-strong); color: var(--paper); }

.bolu-status {
  font-family: var(--font-mono); font-size: 12px;
  color: var(--ink-faint); margin-bottom: 10px;
}

.bolu-matches {
  max-height: 360px; overflow-y: auto;
  border: 1px solid var(--rule); margin-bottom: 18px;
}
.bolu-match-row {
  display: grid;
  grid-template-columns: 1fr 100px 60px 1fr 1fr 110px;
  gap: 10px; padding: 8px 12px;
  border-bottom: 1px solid var(--rule);
  cursor: pointer; font-size: 13px; font-family: var(--font-sans);
}
.bolu-match-row:last-child { border-bottom: 0; }
.bolu-match-row:hover { background: var(--paper-warm); }
.bolu-match-row .col-mono { font-family: var(--font-mono); font-size: 12px; color: var(--ink-soft); }
.bolu-match-row .col-name { font-weight: 500; }

/* BO Card */
.bolu-card {
  background: var(--paper); border: 1px solid var(--rule);
  border-top: 3px solid var(--accent);
  margin-bottom: 16px;
}
.bolu-card-head {
  padding: 14px 18px; border-bottom: 1px solid var(--rule);
  background: var(--paper-warm);
}
.bolu-breadcrumb {
  font-family: var(--font-mono); font-size: 12px;
  color: var(--ink-soft); letter-spacing: 0.06em;
  margin-bottom: 6px;
}
.bolu-breadcrumb .sep { color: var(--ink-faint); margin: 0 6px; }
.bolu-office-name {
  font-family: var(--font-serif); font-size: 26px;
  color: var(--ink-strong); font-weight: 500; line-height: 1.1;
  font-variation-settings: "opsz" 32;
}
.bolu-office-meta {
  font-family: var(--font-mono); font-size: 12px;
  color: var(--ink-faint); margin-top: 6px;
  display: flex; gap: 18px; flex-wrap: wrap;
}
.bolu-office-meta .item strong { color: var(--ink-soft); }

.bolu-subtabs {
  display: flex; gap: 0; padding: 0 18px;
  border-bottom: 1px solid var(--rule); background: var(--paper);
}
.bolu-subtab {
  font-family: var(--font-mono); font-size: 12px;
  padding: 10px 18px; cursor: pointer;
  letter-spacing: 0.10em; text-transform: uppercase;
  color: var(--ink-soft); border: 0; background: transparent;
  border-bottom: 2px solid transparent;
}
.bolu-subtab.active {
  color: var(--ink-strong); border-bottom-color: var(--accent);
}
.bolu-subtab:hover { color: var(--ink); }

.bolu-pane { padding: 18px; display: none; }
.bolu-pane.active { display: block; }

.bolu-table { width: 100%; border-collapse: collapse; font-size: 13px; }
.bolu-table th {
  font-family: var(--font-mono); font-size: 11px;
  letter-spacing: 0.10em; text-transform: uppercase;
  color: var(--ink-soft); font-weight: 500; text-align: right;
  padding: 6px 10px; border-bottom: 1px solid var(--rule);
  background: var(--paper-warm);
}
.bolu-table th:first-child, .bolu-table td:first-child { text-align: left; }
.bolu-table td {
  padding: 6px 10px; border-bottom: 1px solid var(--rule);
  text-align: right; font-family: var(--font-sans);
}
.bolu-table tfoot td {
  font-weight: 600; background: var(--paper-warm);
  border-top: 2px solid var(--rule-strong); border-bottom: 0;
}
.bolu-empty {
  font-family: var(--font-mono); font-size: 13px;
  color: var(--ink-faint); padding: 14px;
  background: var(--paper-warm); border-left: 3px solid var(--rule);
}

.bolu-kpi-strip {
  display: grid; gap: 1px;
  background: var(--rule); border: 1px solid var(--rule);
  margin-bottom: 0;
}
.bolu-kpi-strip.cols-3 { grid-template-columns: repeat(3,1fr); }
.bolu-kpi-strip.cols-4 { grid-template-columns: repeat(4,1fr); }
.bolu-kpi-cell {
  background: var(--paper); padding: 14px 18px;
  display: flex; flex-direction: column; gap: 4px;
}
.bolu-kpi-cell .lbl {
  font-family: var(--font-mono); font-size: 12px;
  letter-spacing: 0.12em; text-transform: uppercase; color: var(--ink-soft);
}
.bolu-kpi-cell .val {
  font-family: var(--font-serif); font-size: 26px;
  font-weight: 500; color: var(--ink-strong); line-height: 1;
  font-variation-settings: "opsz" 32; letter-spacing: -0.02em;
}
.bolu-kpi-cell .val.pos { color: var(--green); }
.bolu-kpi-cell .val.neg { color: var(--red); }
.bolu-kpi-cell .val.nil { color: var(--ink-faint); }
.bolu-kpi-cell .sub {
  font-family: var(--font-mono); font-size: 11.5px;
  color: var(--ink-faint); letter-spacing: 0.04em;
}
.bolu-section-h {
  font-family: var(--font-mono); font-size: 12px;
  letter-spacing: 0.14em; text-transform: uppercase;
  color: var(--ink-soft); margin: 18px 0 8px;
  padding-bottom: 5px; border-bottom: 1px solid var(--rule);
}
/* === END BO LOOKUP STYLES === */
"""

NEW_JS = r"""
// === BO LOOKUP RENDER (auto-generated) ===
(function(){
  var STATE = { query: "", region: "", division: "", currentId: null, currentSub: "booking" };
  var DEBOUNCE = null;

  function nrm(s) { return (s||"").toString().toLowerCase().trim().replace(/\s+/g," "); }
  function fmtN(n) { return (n || 0).toLocaleString("en-IN"); }
  function fmtMoney(n) {
    n = +n || 0;
    if (n >= 1e7) return "₹" + (n/1e7).toFixed(2) + " Cr";
    if (n >= 1e5) return "₹" + (n/1e5).toFixed(2) + " L";
    return "₹" + Math.round(n).toLocaleString("en-IN");
  }

  function applyFilters() {
    var q = nrm(STATE.query);
    var reg = STATE.region, div = STATE.division;
    var res = [];
    for (var i = 0; i < BO_SEARCH.length; i++) {
      var r = BO_SEARCH[i];
      if (reg && r.r !== reg) continue;
      if (div && r.d !== div) continue;
      if (q) {
        if (r.k.indexOf(q) === -1) continue;
      }
      res.push(r);
      if (res.length > 200) break;   // cap
    }
    return res;
  }

  function populateFilters() {
    var regSel = document.getElementById("bolu-region");
    var divSel = document.getElementById("bolu-division");
    if (!regSel || regSel.options.length > 1) return;
    var regs = {}, divs = {};
    BO_SEARCH.forEach(function(r){ regs[r.r]=1; divs[r.d]=1; });
    Object.keys(regs).sort().forEach(function(rg){
      var o = document.createElement("option"); o.value = rg; o.textContent = rg; regSel.appendChild(o);
    });
    Object.keys(divs).sort().forEach(function(dv){
      var o = document.createElement("option"); o.value = dv; o.textContent = dv; divSel.appendChild(o);
    });
  }

  function renderMatches() {
    var matches = applyFilters();
    var status = document.getElementById("bolu-status");
    var box    = document.getElementById("bolu-matches");
    var card   = document.getElementById("bolu-card");

    if (!STATE.query && !STATE.region && !STATE.division) {
      status.textContent = "Start typing an office name, ID, or BO code — or pick a Region/Division.";
      box.innerHTML = ""; box.style.display = "none";
      card.innerHTML = "";
      return;
    }

    if (matches.length === 0) {
      status.textContent = "No offices match your search.";
      box.innerHTML = ""; box.style.display = "none";
      card.innerHTML = "";
      return;
    }
    if (matches.length === 1) {
      status.textContent = "1 office matched.";
      box.style.display = "none";
      box.innerHTML = "";
      showBOCard(matches[0].id);
      return;
    }
    status.textContent = matches.length + " offices matched"
      + (matches.length >= 200 ? " (showing first 200 — narrow your search)" : "")
      + ". Click a row to open.";
    var html = '<div class="bolu-match-row" style="background:var(--paper-warm);font-weight:600;">'
             + '<div>Office Name</div><div>Code</div><div>Type</div>'
             + '<div>Sub-Division</div><div>Division</div><div>Region</div>'
             + '</div>';
    matches.forEach(function(r){
      html += '<div class="bolu-match-row" onclick="boluSelect(\'' + r.id + '\')">'
            + '<div class="col-name">' + r.n + '</div>'
            + '<div class="col-mono">' + (r.c||"—") + '</div>'
            + '<div class="col-mono">' + (r.t||"—") + '</div>'
            + '<div>' + (r.sd||"—") + '</div>'
            + '<div>' + (r.d||"—") + '</div>'
            + '<div>' + (r.r||"—") + '</div>'
            + '</div>';
    });
    box.innerHTML = html;
    box.style.display = "";
    card.innerHTML = "";
  }

  function showBOCard(id) {
    STATE.currentId = id;
    var r = BO_LOOKUP[id];
    var card = document.getElementById("bolu-card");
    if (!r) { card.innerHTML = '<div class="bolu-empty">Office not found.</div>'; return; }

    var bk = r.booking || {total_articles:0,total_amount:0,products:[]};
    var posb = r.posb || {opened:0,closed:0,net:0};
    var pli = r.pli || {policies:0,premium:0};
    var rpli = r.rpli || {policies:0,premium:0};

    var head =
      '<div class="bolu-card-head">'
      + '<div class="bolu-breadcrumb">'
        + (r.region || "—") + '<span class="sep">›</span>'
        + (r.division || "—") + ' Division<span class="sep">›</span>'
        + (r.sub_division || "—")
      + '</div>'
      + '<div class="bolu-office-name">' + r.name + '</div>'
      + '<div class="bolu-office-meta">'
        + '<div class="item"><strong>Office ID:</strong> ' + r.id + '</div>'
        + '<div class="item"><strong>BO Code:</strong> ' + (r.code||"—") + '</div>'
        + '<div class="item"><strong>Type:</strong> ' + (r.type||"—") + '</div>'
        + (r.pincode ? '<div class="item"><strong>Pincode:</strong> ' + r.pincode + '</div>' : '')
      + '</div>'
      + '</div>';

    var tabs =
      '<div class="bolu-subtabs">'
      + '<button class="bolu-subtab" data-sub="booking" onclick="boluSub(\'booking\')">Article Booking</button>'
      + '<button class="bolu-subtab" data-sub="posb" onclick="boluSub(\'posb\')">POSB</button>'
      + '<button class="bolu-subtab" data-sub="ins"  onclick="boluSub(\'ins\')">Insurance</button>'
      + '</div>';

    // Booking pane
    var bkPane;
    if (!bk.products.length) {
      bkPane = '<div class="bolu-empty">No booking activity recorded in Apr–May 2026.</div>';
    } else {
      var rows = "";
      bk.products.forEach(function(p){
        rows += '<tr><td>'+p.n+'</td><td>'+fmtN(p.a)+'</td><td>'+fmtMoney(p.v)+'</td></tr>';
      });
      bkPane =
        '<table class="bolu-table">'
        + '<thead><tr><th>Product</th><th>Articles</th><th>Amount</th></tr></thead>'
        + '<tbody>' + rows + '</tbody>'
        + '<tfoot><tr><td>Total — Apr+May 2026</td><td>'+fmtN(bk.total_articles)+'</td><td>'+fmtMoney(bk.total_amount)+'</td></tr></tfoot>'
        + '</table>';
    }

    // POSB pane
    var netCls = posb.net > 0 ? "pos" : posb.net < 0 ? "neg" : "nil";
    var posbPane =
      '<div class="bolu-kpi-strip cols-3">'
      + '<div class="bolu-kpi-cell"><div class="lbl">Accounts Opened</div><div class="val">'+fmtN(posb.opened)+'</div><div class="sub">In Apr–May 2026</div></div>'
      + '<div class="bolu-kpi-cell"><div class="lbl">Accounts Closed</div><div class="val">'+fmtN(posb.closed)+'</div><div class="sub">In Apr–May 2026</div></div>'
      + '<div class="bolu-kpi-cell"><div class="lbl">Net Addition</div><div class="val '+netCls+'">'+(posb.net>0?'+':'')+fmtN(posb.net)+'</div><div class="sub">'+(posb.net>0?'Positive ✓':posb.net<0?'Negative ✗':'Flat')+'</div></div>'
      + '</div>';

    // Insurance pane
    var nilPli  = pli.policies === 0;
    var nilRpli = rpli.policies === 0;
    var insPane =
      '<div class="bolu-section-h">PLI — Postal Life Insurance</div>'
      + '<div class="bolu-kpi-strip cols-2">'
      + '<div class="bolu-kpi-cell"><div class="lbl">Policies Procured</div><div class="val '+(nilPli?"nil":"")+'">'+fmtN(pli.policies)+'</div><div class="sub">'+(nilPli?'NIL — no PLI procurement':'In Apr–May 2026')+'</div></div>'
      + '<div class="bolu-kpi-cell"><div class="lbl">Premium Collected</div><div class="val '+(nilPli?"nil":"")+'">'+fmtMoney(pli.premium)+'</div><div class="sub">'+(nilPli?'NIL':'₹'+fmtN(pli.policies>0?Math.round(pli.premium/pli.policies):0)+' avg/policy')+'</div></div>'
      + '</div>'
      + '<div class="bolu-section-h">RPLI — Rural Postal Life Insurance</div>'
      + '<div class="bolu-kpi-strip cols-2">'
      + '<div class="bolu-kpi-cell"><div class="lbl">Policies Procured</div><div class="val '+(nilRpli?"nil":"")+'">'+fmtN(rpli.policies)+'</div><div class="sub">'+(nilRpli?'NIL — no RPLI procurement':'In Apr–May 2026')+'</div></div>'
      + '<div class="bolu-kpi-cell"><div class="lbl">Premium Collected</div><div class="val '+(nilRpli?"nil":"")+'">'+fmtMoney(rpli.premium)+'</div><div class="sub">'+(nilRpli?'NIL':'₹'+fmtN(rpli.policies>0?Math.round(rpli.premium/rpli.policies):0)+' avg/policy')+'</div></div>'
      + '</div>';

    var panes =
      '<div id="bolu-pane-booking" class="bolu-pane">'+bkPane+'</div>'
      + '<div id="bolu-pane-posb" class="bolu-pane">'+posbPane+'</div>'
      + '<div id="bolu-pane-ins"  class="bolu-pane">'+insPane+'</div>';

    card.innerHTML = '<div class="bolu-card">' + head + tabs + panes + '</div>';
    boluSub(STATE.currentSub || "booking");
  }

  window.boluSelect = function(id) {
    STATE.currentSub = "booking";
    showBOCard(id);
    document.getElementById("bolu-matches").style.display = "none";
    var card = document.getElementById("bolu-card");
    if (card) card.scrollIntoView({behavior:"smooth", block:"start"});
  };

  window.boluSub = function(sub) {
    STATE.currentSub = sub;
    document.querySelectorAll(".bolu-subtab").forEach(function(b){
      b.classList.toggle("active", b.dataset.sub === sub);
    });
    document.querySelectorAll(".bolu-pane").forEach(function(p){
      p.classList.toggle("active", p.id === "bolu-pane-" + sub);
    });
  };

  function onSearchInput() {
    if (DEBOUNCE) clearTimeout(DEBOUNCE);
    DEBOUNCE = setTimeout(function(){
      STATE.query = document.getElementById("bolu-search").value || "";
      renderMatches();
    }, 120);
  }

  window.boluClear = function() {
    STATE.query = ""; STATE.region = ""; STATE.division = "";
    document.getElementById("bolu-search").value = "";
    document.getElementById("bolu-region").value = "";
    document.getElementById("bolu-division").value = "";
    document.getElementById("bolu-card").innerHTML = "";
    renderMatches();
  };

  window.initBOLookup = function() {
    populateFilters();
    var sb = document.getElementById("bolu-search");
    if (sb && !sb.dataset.bound) {
      sb.addEventListener("input", onSearchInput);
      sb.dataset.bound = "1";
    }
    var rg = document.getElementById("bolu-region");
    var dv = document.getElementById("bolu-division");
    if (rg && !rg.dataset.bound) {
      rg.addEventListener("change", function(){ STATE.region = rg.value; renderMatches(); });
      rg.dataset.bound = "1";
    }
    if (dv && !dv.dataset.bound) {
      dv.addEventListener("change", function(){ STATE.division = dv.value; renderMatches(); });
      dv.dataset.bound = "1";
    }
    renderMatches();
  };
})();
// === END BO LOOKUP RENDER ===
"""

NEW_TAB_BTN = '<button class="tab" onclick="showTab(\'bolookup\', this); initBOLookup();">BO Lookup</button>'

NEW_TAB_HTML = """
<!-- === BO LOOKUP TAB === -->
<div id="tab-bolookup" class="tab-content">
  <div class="section-head">
    <h2>BO Lookup — Branch Office Search</h2>
    <span class="desc">Search by office name, ID, or BO code<br>Across all 10,000+ offices in AP Circle</span>
  </div>
  <div class="bolu-search-bar">
    <div class="bolu-field">
      <label for="bolu-search">Search</label>
      <input id="bolu-search" type="text" placeholder="e.g. Anantapur, 11106440, G5011…" autocomplete="off" />
    </div>
    <div class="bolu-field">
      <label for="bolu-region">Region</label>
      <select id="bolu-region"><option value="">All Regions</option></select>
    </div>
    <div class="bolu-field">
      <label for="bolu-division">Division</label>
      <select id="bolu-division"><option value="">All Divisions</option></select>
    </div>
    <button class="bolu-clear" onclick="boluClear()">Clear</button>
  </div>
  <div id="bolu-status" class="bolu-status">Start typing an office name, ID, or BO code — or pick a Region/Division.</div>
  <div id="bolu-matches" class="bolu-matches" style="display:none;"></div>
  <div id="bolu-card"></div>
</div>
<!-- === END BO LOOKUP TAB === -->
"""


def patch_html(lookup, search_idx):
    html = HTML.read_text(encoding="utf-8")
    print("  Read index.html ({:,} chars)".format(len(html)))

    # 1) DATA constants (inject before SUBDIV_DATA)
    data_block = (
        f"\n{MARK['data_b']}\n"
        f"const BO_LOOKUP = {json.dumps(lookup, ensure_ascii=False, separators=(',', ':'))};\n"
        f"const BO_SEARCH = {json.dumps(search_idx, ensure_ascii=False, separators=(',', ':'))};\n"
        f"{MARK['data_e']}\n"
    )
    if MARK["data_b"] in html:
        pat = re.compile(re.escape(MARK["data_b"]) + r".*?" + re.escape(MARK["data_e"]) + r"\n", re.DOTALL)
        html = pat.sub(data_block.lstrip("\n"), html, count=1); print("  Replaced BO_LOOKUP data")
    else:
        anc = "// === SUBDIV_DATA"
        i = html.find(anc)
        if i < 0:
            anc = "const PLI_POLICIES"; i = html.find(anc)
        if i < 0: print("  ERROR: no anchor"); return False
        html = html[:i] + data_block + html[i:]
        print("  Inserted BO_LOOKUP data")

    # 2) CSS
    if MARK["css_b"] in html:
        pat = re.compile(re.escape(MARK["css_b"]) + r".*?" + re.escape(MARK["css_e"]), re.DOTALL)
        html = pat.sub(NEW_CSS.strip(), html, count=1); print("  Replaced BO Lookup CSS")
    else:
        i = html.rfind("</style>")
        html = html[:i] + NEW_CSS + "\n" + html[i:]; print("  Inserted BO Lookup CSS")

    # 3) JS
    if MARK["js_b"] in html:
        pat = re.compile(re.escape(MARK["js_b"]) + r".*?" + re.escape(MARK["js_e"]) + r"\n", re.DOTALL)
        html = pat.sub(NEW_JS.strip("\n") + "\n", html, count=1); print("  Replaced BO Lookup JS")
    else:
        m = re.search(r"\nfunction showTab\(", html)
        if not m: print("  ERROR: showTab not found"); return False
        html = html[:m.start()] + "\n" + NEW_JS + "\n" + html[m.start():]
        print("  Inserted BO Lookup JS")

    # 4) Tab button — insert AFTER the Cost & ECR button
    cost_btn = "<button class=\"tab\" onclick=\"showTab('cost', this)\">Cost &amp; ECR</button>"
    if "showTab('bolookup'" not in html:
        if cost_btn in html:
            html = html.replace(cost_btn, cost_btn + "\n  " + NEW_TAB_BTN, 1)
            print("  Inserted BO Lookup tab button")
        else:
            print("  WARNING: Cost & ECR tab button not found")

    # 5) Tab HTML — insert after the closing of tab-cost using balanced div counting
    if MARK["tab_b"] in html:
        pat = re.compile(re.escape(MARK["tab_b"]) + r".*?" + re.escape(MARK["tab_e"]), re.DOTALL)
        html = pat.sub(NEW_TAB_HTML.strip(), html, count=1); print("  Replaced BO Lookup tab HTML")
    else:
        i = html.find('<div id="tab-cost"')
        if i < 0: print("  WARNING: tab-cost not found"); return False
        depth, j, end = 0, i, -1
        while j < len(html):
            no = html.find("<div", j); nc = html.find("</div>", j)
            if nc == -1: break
            if no != -1 and no < nc:
                depth += 1; j = no + 4
            else:
                depth -= 1; j = nc + 6
                if depth == 0: end = j; break
        if end < 0: print("  WARNING: tab-cost end not found"); return False
        html = html[:end] + "\n\n" + NEW_TAB_HTML.strip() + "\n" + html[end:]
        print("  Inserted BO Lookup tab HTML")

    HTML.write_text(html, encoding="utf-8")
    print("  Wrote index.html ({:,} chars)".format(len(html)))
    return True


if __name__ == "__main__":
    print("=== patch_bo_lookup.py ===")
    lookup, name_map = read_posb()
    merge_drilldown(PLI_XLSX,  "pli",  lookup, name_map)
    merge_drilldown(RPLI_XLSX, "rpli", lookup, name_map)
    merge_booking(lookup)
    search_idx = build_search_index(lookup)
    print(f"\n  BO_LOOKUP: {len(lookup):,} offices")
    print(f"  BO_SEARCH: {len(search_idx):,} entries")
    # Sample
    sample_id = next(iter(lookup))
    s = lookup[sample_id]
    print(f"  Sample [{sample_id}]: {s['name']} / posb_net={s['posb']['net']} / "
          f"pli={s['pli']['policies']} / rpli={s['rpli']['policies']} / "
          f"booking_articles={s['booking']['total_articles']} ({len(s['booking']['products'])} products)")
    print("\nPatching index.html...")
    if patch_html(lookup, search_idx):
        print("\n=== Done ===")

#!/usr/bin/env python3
"""
patch_june2026_circle_full.py
────────────────────────────────
Recomputes CIRCLE_FULL (the "AP Circle - consolidated position" entire-circle
revenue/expenditure/ECR figures, introduced in commit 394821d) from the June
ECR file. This constant is separate from DATA_BY_MONTH and was NOT updated by
patch_june2026_ecr.py, so it was still showing May's totals.

Formula (reverse-engineered from May's committed values, verified to
reproduce them exactly):
  CIRCLE_FULL = sum over the 33 divisions/RMS units + 6 admin/regional/PSD
  rows (AP GM (Finance), Andhra Pradesh Circle, PSD Vijayawada, Kurnool
  Region, Vijayawada Region, Visakhapatnam Region) that DATA_BY_MONTH's
  33-entity sum drops.

Run:  python patch_june2026_circle_full.py
"""
import json, re
from pathlib import Path
import openpyxl

BASE = Path(__file__).parent
HTML = BASE / "index.html"
XLSX = BASE / "uploads" / "june_2026" / "ECR_Calculation_June2026.xlsx"
CR = 10_000_000

DIVISIONS = {
    "Amalapuram", "Anakapalle", "Anantapur", "Bhimavaram", "Chittoor",
    "Cuddapah", "Eluru", "Gudivada", "Gudur", "Guntur", "Hindupur",
    "Kakinada", "Kurnool", "Machilipatnam", "Markapur", "Nandyal",
    "Narasaraopet", "Nellore", "Parvathipuram", "Prakasam", "Proddatur",
    "RMS AG", "RMS TP", "RMS V", "RMS Y", "Rajahmundry", "Srikakulam",
    "Tadepalligudem", "Tenali", "Tirupati", "Vijayawada", "Visakhapatnam",
    "Vizianagaram",
}
ADMIN_ROWS = {
    "AP GM (Finance)", "Andhra Pradesh Circle", "PSD Vijayawada",
    "Kurnool Region", "Vijayawada Region", "Visakhapatnam Region",
}
ALL_ROWS = DIVISIONS | ADMIN_ROWS

COL_DIV, COL_POSB, COL_PLIR = 0, 8, 24
COL_PARC, COL_MO, COL_IRGB, COL_CCS = 25, 26, 27, 28
COL_TOT_R, COL_TOT_E = 31, 34

# SWAP sheet column order is NOT stable between monthly files — look up by
# header name instead of hardcoded index (June reordered/renamed columns vs May).
SWAP_HEADERS = {
    "Salaries": "Salaries", "Wages": "Wages", "Allowances": "Allowances", "Pension": "Pension",
    "Plan": "Plan Expnditure ",
}
SWAP_OTHER_CANDIDATES = ["Others", "Others than swap"]


def to_cr(v):
    return (v or 0) / CR


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ecr_ws, swap_ws = wb["ECR"], wb["SWAP"]

    actuals = {"Parcel": 0.0, "Mail Operations": 0.0, "IR & GB": 0.0, "CCS": 0.0,
               "POSB": 0.0, "PLI/RPLI": 0.0, "Total": 0.0}
    total_exp = 0.0
    seen = set()
    for row in ecr_ws.iter_rows(min_row=2, values_only=True):
        name = row[COL_DIV]
        if not name or name not in ALL_ROWS:
            continue
        seen.add(name)
        actuals["Parcel"] += to_cr(row[COL_PARC])
        actuals["Mail Operations"] += to_cr(row[COL_MO])
        actuals["IR & GB"] += to_cr(row[COL_IRGB])
        actuals["CCS"] += to_cr(row[COL_CCS])
        actuals["POSB"] += to_cr(row[COL_POSB])
        actuals["PLI/RPLI"] += to_cr(row[COL_PLIR])
        actuals["Total"] += to_cr(row[COL_TOT_R])
        total_exp += to_cr(row[COL_TOT_E])

    missing = ALL_ROWS - seen
    if missing:
        print(f"WARNING: rows not found in ECR sheet: {sorted(missing)}")

    swap_headers = [c.value for c in next(swap_ws.iter_rows(min_row=1, max_row=1))]
    other_header = next((h for h in SWAP_OTHER_CANDIDATES if h in swap_headers), None)
    if other_header is None:
        raise SystemExit(f"ERROR: no 'Others' column found in SWAP sheet headers: {swap_headers}")
    col_idx = {label: swap_headers.index(hdr) for label, hdr in SWAP_HEADERS.items()}
    col_idx["Other"] = swap_headers.index(other_header)

    swap = {"Salaries": 0.0, "Wages": 0.0, "Pension": 0.0, "Allowances": 0.0, "Plan": 0.0, "Other": 0.0}
    for row in swap_ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        if not name or name not in ALL_ROWS:
            continue
        swap["Salaries"] += to_cr(row[col_idx["Salaries"]])
        swap["Wages"] += to_cr(row[col_idx["Wages"]])
        swap["Pension"] += to_cr(row[col_idx["Pension"]])
        swap["Allowances"] += to_cr(row[col_idx["Allowances"]])
        swap["Plan"] += to_cr(row[col_idx["Plan"]])
        swap["Other"] += to_cr(row[col_idx["Other"]])
    swap["Total"] = sum(swap.values())

    ecr_with_pension = (actuals["Total"] / total_exp * 100) if total_exp > 0 else 0
    ecr_without_pension = (actuals["Total"] / (total_exp - swap["Pension"]) * 100) if (total_exp - swap["Pension"]) > 0 else 0

    circle_full = {
        "actuals": {k: round(v, 4) for k, v in actuals.items()},
        "total_exp": round(total_exp, 4),
        "swap": {k: round(v, 4) for k, v in swap.items()},
        "ecr_with_pension": round(ecr_with_pension, 4),
        "ecr_without_pension": round(ecr_without_pension, 4),
    }
    print("New CIRCLE_FULL:", json.dumps(circle_full, indent=2))

    html = HTML.read_text(encoding="utf-8")
    start = html.find("const CIRCLE_FULL")
    if start < 0:
        raise SystemExit("ERROR: const CIRCLE_FULL not found")
    end = html.find("};", start) + 2
    new_const = "const CIRCLE_FULL = " + json.dumps(circle_full, ensure_ascii=False) + ";"
    html = html[:start] + new_const + html[end:]
    HTML.write_text(html, encoding="utf-8")
    print(f"\nReplaced CIRCLE_FULL in index.html")


if __name__ == "__main__":
    main()

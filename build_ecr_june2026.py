#!/usr/bin/env python3
"""
build_ecr_june2026.py
──────────────────────
Reads uploads/june_2026/ECR_Calculation_June2026.xlsx (same column layout as
May) and writes _ecr_june2026.json with cumulative (Apr+May+Jun) revenue,
expenditure, and SWAP breakdown per division — same shape as the May patch
used to inject into DATA_BY_MONTH['May-26'].

Run:  python build_ecr_june2026.py
"""
import json
from pathlib import Path
import openpyxl

BASE = Path(__file__).parent
XLSX = BASE / "uploads" / "june_2026" / "ECR_Calculation_June2026.xlsx"
CR = 10_000_000

DIVISIONS = {
    "Amalapuram", "Anakapalle", "Anantapur", "Bhimavaram", "Chittoor",
    "Cuddapah", "Eluru", "Gudivada", "Gudur", "Guntur", "Hindupur",
    "Kakinada", "Kurnool", "Machilipatnam", "Markapur", "Nandyal",
    "Narasaraopet", "Nellore", "Parvathipuram", "Prakasam", "Proddatur",
    "RMS AG", "RMS TP", "RMS V", "RMS Y", "Rajahmundry", "Srikakulam",
    "Tadepalligudem", "Tenali", "Tirupati", "Vijayawada", "Visakhapatnam",
    "Vizianagaram",
}

COL_DIV   = 0
COL_POSB  = 8   # Total rev on sbcc up to June-2026
COL_PLIR  = 24  # Total PLI/RPLI Rev.
COL_PARC  = 25  # Parcel
COL_MO    = 26  # MO
COL_IRGB  = 27  # IRGB
COL_CCS   = 28  # CCS
COL_TOT_R = 31  # Total revenue
COL_TOT_E = 34  # Exp. For ECR calculation
COL_PEN   = 36  # 320107 (pension component)

# SWAP sheet column order is NOT stable between monthly files (June re-ordered
# and renamed columns vs May) — look columns up by header name instead of index.
SWAP_HEADERS = {
    "Salaries": "Salaries", "Wages": "Wages", "Allowances": "Allowances",
    "Pension": "Pension", "Plan": "Plan Expnditure ", "Total": "Total expenditure",
}
SWAP_OTHER_CANDIDATES = ["Others", "Others than swap"]


def to_cr(v):
    return round((v or 0) / CR, 7)


def read_ecr():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ecr_ws = wb["ECR"]
    swap_ws = wb["SWAP"]
    out = {}

    for row in ecr_ws.iter_rows(min_row=2, values_only=True):
        name = row[COL_DIV]
        if not name or name not in DIVISIONS:
            continue
        out[name] = {
            "actuals": {
                "Parcel":          to_cr(row[COL_PARC]),
                "Mail Operations": to_cr(row[COL_MO]),
                "IR & GB":         to_cr(row[COL_IRGB]),
                "CCS":             to_cr(row[COL_CCS]),
                "POSB":            to_cr(row[COL_POSB]),
                "PLI/RPLI":        to_cr(row[COL_PLIR]),
                "Total":           to_cr(row[COL_TOT_R]),
            },
            "total_exp": to_cr(row[COL_TOT_E]),
        }

    swap_headers = [c.value for c in next(swap_ws.iter_rows(min_row=1, max_row=1))]
    other_header = next((h for h in SWAP_OTHER_CANDIDATES if h in swap_headers), None)
    if other_header is None:
        raise SystemExit(f"ERROR: no 'Others' column found in SWAP sheet headers: {swap_headers}")
    col_idx = {label: swap_headers.index(hdr) for label, hdr in SWAP_HEADERS.items()}
    col_idx["Other"] = swap_headers.index(other_header)

    for row in swap_ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        if not name or name not in DIVISIONS or name not in out:
            continue
        out[name]["swap"] = {
            "Salaries":   to_cr(row[col_idx["Salaries"]]),
            "Wages":      to_cr(row[col_idx["Wages"]]),
            "Pension":    to_cr(row[col_idx["Pension"]]),
            "Allowances": to_cr(row[col_idx["Allowances"]]),
            "Plan":       to_cr(row[col_idx["Plan"]]),
            "Other":      to_cr(row[col_idx["Other"]]),
            "Total":      to_cr(row[col_idx["Total"]]),
        }

    return out


if __name__ == "__main__":
    data = read_ecr()
    missing = DIVISIONS - set(data.keys())
    if missing:
        print(f"WARNING: missing in Excel: {sorted(missing)}")
    (BASE / "_ecr_june2026.json").write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
    print(f"Loaded ECR cumulative (Apr+May+Jun) values for {len(data)} divisions")
    sample = next(iter(data.items()))
    print("Sample:", sample)
